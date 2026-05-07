"""
Regencare Competitive Study Excel Generator
Covers top 10 competitor sites ranked by keyword presence
"""
import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image
import re

wb = openpyxl.Workbook()

# ── COLOURS ───────────────────────────────────────────────────────────────────
NAVY    = "0D2B55"
GOLD    = "C9A02D"
WHITE   = "FFFFFF"
LIGHT   = "F4F6F9"
MID     = "D0D8E4"
TEXT    = "1A1A2E"
RED     = "C0392B"
RED_L   = "FDECEA"
GREEN   = "1A7A4A"
GREEN_L = "E8F5EE"
AMB     = "D68A00"
AMB_L   = "FFF3CD"
TEAL    = "0D6E6E"
TEAL_L  = "E0F4F4"
GREY    = "6C757D"
GREY_L  = "F0F0F0"
CRIT    = "8B0000"   # dark red for critical

def fill(hex_):
    return PatternFill("solid", fgColor=hex_)

def font(hex_=TEXT, bold=False, sz=10, italic=False):
    return Font(color=hex_, bold=bold, size=sz, italic=italic, name="Calibri")

def border(color=MID):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def align(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def set_col(ws, col_idx, width):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

def set_row(ws, row_idx, height):
    ws.row_dimensions[row_idx].height = height

def hdr(ws, row, col, text, bg=NAVY, fg=WHITE, bold=True, sz=10,
        ha="center", wrap=True, border_color=MID):
    c = ws.cell(row=row, column=col, value=text)
    c.fill    = fill(bg)
    c.font    = font(fg, bold=bold, sz=sz)
    c.alignment = align(ha, "center", wrap)
    c.border  = border(border_color)
    return c

def cell(ws, row, col, text, bg=WHITE, fg=TEXT, bold=False, sz=10,
         ha="left", wrap=True):
    c = ws.cell(row=row, column=col, value=text)
    c.fill    = fill(bg)
    c.font    = font(fg, bold=bold, sz=sz)
    c.alignment = align(ha, "center", wrap)
    c.border  = border()
    return c

def score_cell(ws, row, col, score, label=""):
    """Score 1-5 colour-coded cell"""
    colours = {1: RED, 2: "E05C00", 3: AMB, 4: TEAL, 5: GREEN}
    bgs     = {1: RED_L, 2: AMB_L, 3: AMB_L, 4: TEAL_L, 5: GREEN_L}
    c = ws.cell(row=row, column=col, value=f"{score}/5 {label}".strip())
    c.fill = fill(bgs.get(score, WHITE))
    c.font = font(colours.get(score, TEXT), bold=True, sz=10)
    c.alignment = align("center", "center", False)
    c.border = border()
    return c

def tick(ws, row, col, val):
    """✓ / ✗ / ~ cell"""
    cfg = {
        "Yes":      (GREEN_L, GREEN, "✓ Yes"),
        "No":       (RED_L,   RED,   "✗ No"),
        "Partial":  (AMB_L,   AMB,   "~ Partial"),
        "Unknown":  (GREY_L,  GREY,  "? Unknown"),
        "Strong":   (GREEN_L, GREEN, "✓ Strong"),
        "Weak":     (AMB_L,   AMB,   "~ Weak"),
        "None":     (RED_L,   RED,   "✗ None"),
    }
    bg, fg, txt = cfg.get(val, (WHITE, TEXT, val))
    c = ws.cell(row=row, column=col, value=txt)
    c.fill = fill(bg)
    c.font = font(fg, bold=True, sz=10)
    c.alignment = align("center", "center", False)
    c.border = border()
    return c

def pos_cell(ws, row, col, pos_str):
    """Position cell: colour by ranking position band"""
    if pos_str in ("—", "Not Ranking", "NR"):
        bg, fg, val = RED_L, RED, "Not Ranking"
    elif pos_str == "Regencare":
        bg, fg, val = TEAL_L, TEAL, "Regencare"
    else:
        try:
            n = int(re.sub(r'\D','', str(pos_str)))
            if   n <= 3:  bg, fg, val = GREEN_L, GREEN, f"#{n}"
            elif n <= 7:  bg, fg, val = AMB_L,   AMB,   f"#{n}"
            elif n <= 10: bg, fg, val = RED_L,   RED,   f"#{n}"
            else:         bg, fg, val = GREY_L,  GREY,  f"#{n}"
        except:
            bg, fg, val = GREY_L, GREY, str(pos_str)
    c = ws.cell(row=row, column=col, value=val)
    c.fill = fill(bg)
    c.font = font(fg, bold=True, sz=10)
    c.alignment = align("center", "center", False)
    c.border = border()
    return c

def merge_hdr(ws, r1, c1, r2, c2, text, bg=NAVY, fg=WHITE, sz=11, bold=True):
    ws.merge_cells(start_row=r1, start_column=c1,
                   end_row=r2, end_column=c2)
    c = ws.cell(row=r1, column=c1, value=text)
    c.fill = fill(bg)
    c.font = font(fg, bold=bold, sz=sz)
    c.alignment = align("center", "center", True)
    c.border = border()
    return c

# ═════════════════════════════════════════════════════════════════════════════
# MASTER DATA
# ═════════════════════════════════════════════════════════════════════════════

COMPETITORS = [
    # (ID, Name, Domain, Type, Tier, Locations, Focus, Threat)
    (1,  "Orthogen Care",        "orthogencare.com",       "Brand Cannibalization",  "CRITICAL", "Ernakulam/Kochi",                   "Ortho PRP / Regenerative Medicine",      "CRITICAL"),
    (2,  "DermaVue",             "dermavue.com",           "Kerala Clinic Chain",    "HIGH",     "7 clinics, Kerala + Tamil Nadu",    "GFC, PRP, Hair Loss, Trichology",        "HIGH"),
    (3,  "Oliva Clinic",         "olivaclinic.com",        "National Chain",         "HIGH",     "Pan-India (Kochi branch present)",  "Hair Fall, PRP, Hair Transplant",        "HIGH"),
    (4,  "Cutis International",  "cutisinternational.com", "International Chain",    "HIGH",     "Kerala + UAE + UK (10 locations)",  "Hair Transplant, Skin, Plastic Surgery", "HIGH"),
    (5,  "La Densitae",          "ladensitae.com",         "Hair Transplant Chain",  "MED-HIGH", "13+ branches, India + Dubai",       "Hair Transplant FUE/FUT, PRP",           "MEDIUM"),
    (6,  "Zaaya Skin Clinic",    "zaayaskinclinic.com",    "Local Kochi Clinic",     "MED-HIGH", "Aluva, Kakkanad, Kadavanthra",      "GFC, PRP, Hair Loss, Skin",              "MEDIUM"),
    (7,  "DHI International",    "dhiinternational.com",   "Global Hair Brand",      "MEDIUM",   "Global — Kochi branch (Kadavanthr)","Hair Transplant (DHI technique)",        "MEDIUM"),
    (8,  "DH Clinic",            "dh-clinic.com",          "Ortho + Stem Cell",      "MEDIUM",   "Kerala (multi-branch)",             "Ortho, Stem Cell, PRP (Surecell AU)",   "MEDIUM"),
    (9,  "Hair Wellness Clinic", "hairwellnessclinic.com", "Specialized Hair Clinic","MEDIUM",   "Kerala",                            "GFC, PRP, Hair treatments",              "MEDIUM"),
    (10, "Epione Pain Centre",   "paincentre.in",          "Pain & Regen Clinic",    "MEDIUM",   "Kerala",                            "PRP, Stem Cell, Prolotherapy, Pain",     "MEDIUM"),
]

# Top 20 keywords by search volume
KEYWORDS = [
    # (keyword, est_vol, intent, cluster)
    ("hair loss treatment Kochi",           2400, "Commercial", "Hair — Kochi"),
    ("hair transplant Kochi",               1900, "Transact.",  "Hair — Kochi"),
    ("PRP treatment Kerala",                1600, "Commercial", "PRP — Kerala"),
    ("GFC therapy Kochi",                   1300, "Commercial", "GFC — Kochi"),
    ("GFC hair treatment Kochi",            1000, "Transact.",  "GFC — Kochi"),
    ("hair loss clinic Kochi",               900, "Commercial", "Hair — Kochi"),
    ("knee pain treatment without surgery",  720, "Commercial", "Ortho — Kerala"),
    ("regenerative medicine Kerala",         590, "Navigat.",   "Regen — Kerala"),
    ("PRP treatment Kochi",                  590, "Commercial", "PRP — Kerala"),
    ("stem cell therapy Kerala",             480, "Commercial", "Regen — Kerala"),
    ("PRP treatment Chennai",                480, "Commercial", "PRP — Chennai"),
    ("hair transplant Calicut",              390, "Transact.",  "Hair — Calicut"),
    ("hair regrowth treatment Kerala",       390, "Commercial", "Hair — Kerala"),
    ("PRP for knee pain Kochi",              320, "Commercial", "Ortho — Kochi"),
    ("GFC therapy Calicut",                  260, "Commercial", "GFC — Calicut"),
    ("best orthopedic doctor Kochi",         210, "Commercial", "Ortho — Kochi"),
    ("regenerative medicine Kochi",          170, "Navigat.",   "Regen — Kochi"),
    ("GFC therapy cost Kerala",              170, "Transact.",  "Conversational"),
    ("stem cell treatment Kochi",            140, "Commercial", "Regen — Kochi"),
    ("non-surgical knee treatment Kerala",   170, "Commercial", "Ortho — Kerala"),
]

# Keyword positions per competitor (10 comps x 20 keywords)
# Format: string position "1","3","NR" etc.
# Rows = COMPETITORS (10), Cols = KEYWORDS (20)
POSITIONS = {
    # comp_id: [pos for each of 20 keywords]
    1:  ["NR","NR","NR","NR","NR","NR","3","4","NR","3","NR","NR","NR","4","NR","5","4","NR","5","4"],     # orthogencare
    2:  ["2","NR","4","1","2","3","NR","NR","3","NR","NR","NR","NR","NR","NR","NR","NR","NR","NR","NR"],   # dermavue
    3:  ["1","NR","2","NR","NR","2","NR","NR","2","NR","NR","NR","4","NR","NR","NR","NR","NR","NR","NR"],  # oliva
    4:  ["NR","3","NR","7","6","NR","NR","NR","NR","NR","NR","2","NR","NR","5","NR","NR","NR","NR","NR"],  # cutis
    5:  ["4","2","NR","5","4","4","NR","NR","NR","NR","NR","3","NR","NR","NR","NR","NR","NR","NR","NR"],   # la densitae
    6:  ["5","NR","NR","3","3","5","NR","NR","NR","NR","NR","NR","NR","NR","NR","NR","NR","NR","NR","NR"], # zaaya
    7:  ["NR","1","NR","NR","NR","NR","NR","NR","NR","NR","NR","1","NR","NR","NR","NR","NR","NR","NR","NR"],# dhi
    8:  ["NR","NR","NR","NR","NR","NR","5","6","NR","5","NR","NR","NR","5","NR","6","NR","NR","5","6"],    # dh-clinic
    9:  ["NR","NR","NR","8","7","NR","NR","NR","NR","NR","NR","NR","NR","NR","NR","NR","NR","NR","NR","NR"],# hair wellness
    10: ["NR","NR","NR","NR","NR","NR","7","8","NR","7","NR","NR","NR","7","NR","NR","NR","NR","7","8"],   # epione
}
# Regencare's own positions for same keywords
REGENCARE_POS = ["NR","NR","3","9","NR","NR","2","1","4","1","3","NR","NR","3","NR","NR","2","NR","2","2"]

# On-Page Audit data per competitor
# (title_opt, single_h1, meta_opt, schema, faq_page, word_ct, pricing, booking, mobile, nmc_ok)
ONPAGE = {
    1:  ("Strong",  "Yes",     "Partial", "None",    "No",     2000,  "No",      "Yes",  "Unknown","Yes"),
    2:  ("Strong",  "Yes",     "Strong",  "Strong",  "No",     4900,  "Yes",     "Yes",  "Yes",    "Yes"),
    3:  ("Strong",  "Yes",     "Strong",  "Strong",  "Yes",    3200,  "Yes",     "Yes",  "Yes",    "Yes"),
    4:  ("Partial", "Weak",    "Partial", "Partial", "No",     2100,  "No",      "Yes",  "Yes",    "Partial"),
    5:  ("Partial", "Yes",     "Partial", "None",    "No",     1200,  "Partial", "Yes",  "Yes",    "Partial"),
    6:  ("Strong",  "Yes",     "Partial", "None",    "No",     2300,  "No",      "Yes",  "Yes",    "Yes"),
    7:  ("Strong",  "Yes",     "Strong",  "Strong",  "No",     2800,  "No",      "Yes",  "Yes",    "Yes"),
    8:  ("Partial", "Yes",     "Partial", "None",    "No",     1800,  "No",      "Yes",  "Yes",    "Yes"),
    9:  ("Partial", "Unknown", "Partial", "None",    "No",     1400,  "Partial", "Yes",  "Unknown","Yes"),
    10: ("Partial", "Unknown", "Partial", "None",    "No",     1600,  "No",      "Yes",  "Unknown","Yes"),
}

# Authority & Trust
# (google_rating, review_count, years, da_est, backlinks_est, doctor_credentials, content_authority)
AUTHORITY = {
    1:  (None,  None,  "Unknown", 22, "Low",    "Dr. Vineeth MB — MBBS, MS Ortho",            "Medium"),
    2:  (4.8,   1438,  "8+",      34, "Medium", "Dr. Minu Liz Mathew — MBBS, MD DVL",         "High"),
    3:  (4.9,   1475,  "15+",     52, "High",   "4 x MD Dermatologists (Kochi branch)",        "Very High"),
    4:  (4.7,   500,   "10+",     38, "Medium", "Dr. Abraham Thomas + 6 specialists",          "Medium"),
    5:  (4.5,   None,  "12+",     29, "Low",    "No named doctors on page",                   "Low"),
    6:  (4.6,   None,  "5+",      18, "Low",    "Transfusion Medicine Consultant (unnamed)",   "Medium"),
    7:  (4.8,   None,  "20+",     55, "High",   "Not listed on landing page",                  "High"),
    8:  (None,  None,  "Unknown", 16, "Low",    "Dr. Althaaf Mohamed A.H. — Arthroscopy",     "Low"),
    9:  (None,  None,  "Unknown", 12, "Low",    "Not listed",                                  "Low"),
    10: (None,  None,  "7+",      19, "Low",    "Not listed",                                  "Low"),
}

# Service Coverage (Yes/No/Partial per service per competitor)
SERVICES = [
    "GFC Therapy",
    "PRP — Hair",
    "PRP — Ortho / Joints",
    "Stem Cell Therapy",
    "Hair Transplant (FUE)",
    "Platelet / Exosome Therapy",
    "Skin / Dermatology",
    "Sports Injury Treatment",
    "Knee Pain (non-surgical)",
    "Spine / Back Pain",
    "UAE / Medical Tourism Page",
    "Doctor Profile Pages",
    "Blog / Research Content",
    "Online Booking (real)",
    "Cost / Pricing Page",
]
SVC_DATA = {
    #  GFC  PRP-H PRP-O  SC   FUE  EXO  SKIN SPORT KNEE SPINE UAE  DOC  BLOG BOOK COST
    1: ["No","No","Yes","Partial","No","No","No","Yes","Yes","Yes","No","Yes","Yes","Yes","No"],
    2: ["Yes","Yes","No","No","Yes","No","Yes","No","No","No","No","Yes","Yes","Yes","Yes"],
    3: ["No","Yes","No","No","Yes","Yes","Yes","No","No","No","No","Yes","Yes","Yes","Yes"],
    4: ["Yes","Yes","No","No","Yes","Yes","Yes","No","No","No","Yes","Yes","No","Yes","No"],
    5: ["No","Yes","No","No","Yes","No","No","No","No","No","Yes","No","Partial","Yes","Partial"],
    6: ["Yes","Yes","No","No","No","No","Yes","No","No","No","No","No","Yes","Yes","No"],
    7: ["No","No","No","No","Yes","No","No","No","No","No","Yes","No","No","Yes","No"],
    8: ["No","No","Yes","Yes","No","No","No","Yes","Yes","Yes","No","Yes","Partial","Yes","No"],
    9: ["Yes","Yes","No","No","No","No","No","No","No","No","No","No","No","Yes","Partial"],
    10:["No","No","Yes","Yes","No","No","No","Yes","Yes","Yes","No","No","Partial","Yes","No"],
}

# ═════════════════════════════════════════════════════════════════════════════
# SHEET 1 — COVER
# ═════════════════════════════════════════════════════════════════════════════
ws_cover = wb.active
ws_cover.title = "COVER"
ws_cover.sheet_view.showGridLines = False

for c in range(1, 12):
    ws_cover.column_dimensions[get_column_letter(c)].width = 14

# Navy header block rows 1-6
for r in range(1, 8):
    for c in range(1, 12):
        ws_cover.cell(r, c).fill = fill(NAVY)
    set_row(ws_cover, r, 22)

# Title
ws_cover.merge_cells("A1:K3")
tc = ws_cover.cell(1, 1, "REGENCARE.IN  —  COMPETITIVE INTELLIGENCE STUDY")
tc.fill = fill(NAVY); tc.font = font(GOLD, True, 18)
tc.alignment = align("center", "center", False)

ws_cover.merge_cells("A4:K5")
sc = ws_cover.cell(4, 1, "Top-10 Competitor Analysis  |  Keyword-Driven  |  May 2026")
sc.fill = fill(NAVY); sc.font = font(WHITE, False, 12, italic=True)
sc.alignment = align("center", "center", False)

ws_cover.merge_cells("A6:K7")
pc = ws_cover.cell(6, 1, "Prepared by NT Global Digital  ·  Client: Regencare.in  ·  Confidential")
pc.fill = fill(NAVY); pc.font = font(MID, False, 10)
pc.alignment = align("center", "center", False)

# Spacer row 8
ws_cover.row_dimensions[8].height = 14

# ── STUDY SCOPE ──
merge_hdr(ws_cover, 9, 1, 9, 11, "STUDY SCOPE & METHODOLOGY", NAVY, WHITE, 12)
scope_items = [
    ("Data Sources", "Live Google SERP captures (incognito, geolocated) · Direct site audits via web fetch · Google Business Profile checks"),
    ("Keywords Analysed", "20 highest-volume keywords across GFC, PRP, Hair, Ortho, Stem Cell, Regenerative Medicine clusters"),
    ("Competitors Tracked", "10 direct SERP competitors identified by frequency of appearance across priority keyword SERPs"),
    ("Dimensions Studied", "Keyword positions · On-page SEO · Content depth · Trust & authority · Service coverage · White space"),
    ("Locations Covered", "Kochi, Calicut, Chennai (Regencare's three branches) + UAE/NRI segment"),
    ("Output",  "8-sheet Excel: Competitor Overview · Keyword Ranking Map · On-Page Audit · Authority · Service Map · Gaps · Actions"),
]
for i, (label, detail) in enumerate(scope_items):
    r = 10 + i
    set_row(ws_cover, r, 28)
    bg = LIGHT if i % 2 == 0 else WHITE
    ws_cover.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    lc = ws_cover.cell(r, 1, label)
    lc.fill = fill(bg); lc.font = font(NAVY, True, 10)
    lc.alignment = align("left", "center"); lc.border = border()
    ws_cover.merge_cells(start_row=r, start_column=4, end_row=r, end_column=11)
    dc = ws_cover.cell(r, 4, detail)
    dc.fill = fill(bg); dc.font = font(TEXT, False, 10)
    dc.alignment = align("left", "center"); dc.border = border()

# Spacer
ws_cover.row_dimensions[16].height = 14

# ── THREAT TIER LEGEND ──
merge_hdr(ws_cover, 17, 1, 17, 11, "THREAT TIER LEGEND", NAVY, WHITE, 11)
legend = [
    ("CRITICAL", RED,   "Same doctor domain directly cannibalizing Regencare SERPs — immediate resolution required"),
    ("HIGH",     AMB,   "National or multi-location clinic actively outranking Regencare on high-volume keywords"),
    ("MEDIUM",   TEAL,  "Local or specialist competitor with partial keyword overlap — monitor and address in content strategy"),
]
for i, (tier, col, desc) in enumerate(legend):
    r = 18 + i
    set_row(ws_cover, r, 24)
    ws_cover.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    tc2 = ws_cover.cell(r, 1, tier)
    tc2.fill = fill(col if col != AMB else "E07B00"); tc2.font = font(WHITE, True, 11)
    tc2.alignment = align("center", "center"); tc2.border = border()
    ws_cover.merge_cells(start_row=r, start_column=3, end_row=r, end_column=11)
    dc2 = ws_cover.cell(r, 3, desc)
    dc2.fill = fill(LIGHT); dc2.font = font(TEXT, False, 10)
    dc2.alignment = align("left", "center"); dc2.border = border()

# Spacer
ws_cover.row_dimensions[21].height = 14

# ── POSITION COLOUR LEGEND ──
merge_hdr(ws_cover, 22, 1, 22, 11, "RANKING POSITION COLOUR LEGEND (Keyword Ranking Map Sheet)", NAVY, WHITE, 11)
pos_legend = [
    ("#1 – #3",       GREEN_L, GREEN, "Top 3 — strong ranking, high traffic likely"),
    ("#4 – #7",       AMB_L,   AMB,   "Mid-page — visible but losing clicks to top 3"),
    ("#8 – #10",      RED_L,   RED,   "Bottom page 1 — minimal traffic share"),
    ("Not Ranking",   RED_L,   RED,   "No presence on page 1 for this keyword"),
]
for i, (band, bg, fg, desc) in enumerate(pos_legend):
    r = 23 + i
    set_row(ws_cover, r, 22)
    ws_cover.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    bc = ws_cover.cell(r, 1, band)
    bc.fill = fill(bg); bc.font = font(fg, True, 10)
    bc.alignment = align("center", "center"); bc.border = border()
    ws_cover.merge_cells(start_row=r, start_column=4, end_row=r, end_column=11)
    dc3 = ws_cover.cell(r, 4, desc)
    dc3.fill = fill(LIGHT); dc3.font = font(TEXT, False, 10)
    dc3.alignment = align("left", "center"); dc3.border = border()

# ── SHEET INDEX ──
ws_cover.row_dimensions[27].height = 14
merge_hdr(ws_cover, 28, 1, 28, 11, "SHEET INDEX", NAVY, WHITE, 11)
sheets_idx = [
    ("1 — COVER",              "This page — study scope, legend, methodology"),
    ("2 — COMPETITOR OVERVIEW","Snapshot of all 10 competitors: domain, type, locations, focus, threat tier"),
    ("3 — KEYWORD RANKING MAP","20 keywords × 10 competitors — who ranks where vs. Regencare"),
    ("4 — ON-PAGE SEO AUDIT",  "Page-level technical and content quality: title, H1, schema, word count, pricing, booking"),
    ("5 — AUTHORITY & TRUST",  "Google rating, review count, DA estimate, doctor credentials, E-E-A-T score"),
    ("6 — SERVICE COVERAGE",   "Which treatments each competitor offers vs. Regencare — gap & overlap matrix"),
    ("7 — WHITE SPACE & GAPS", "Topics, keywords, and SERP features nobody owns well — Regencare's fastest opportunities"),
    ("8 — ACTION PRIORITIES",  "Prioritised action list: what to do first based on competitive study findings"),
]
for i, (name, desc) in enumerate(sheets_idx):
    r = 29 + i
    set_row(ws_cover, r, 22)
    bg = LIGHT if i % 2 == 0 else WHITE
    ws_cover.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    nc = ws_cover.cell(r, 1, name)
    nc.fill = fill(bg); nc.font = font(NAVY, True, 10)
    nc.alignment = align("left", "center"); nc.border = border()
    ws_cover.merge_cells(start_row=r, start_column=4, end_row=r, end_column=11)
    dc4 = ws_cover.cell(r, 4, desc)
    dc4.fill = fill(bg); dc4.font = font(TEXT, False, 10)
    dc4.alignment = align("left", "center"); dc4.border = border()

# ═════════════════════════════════════════════════════════════════════════════
# SHEET 2 — COMPETITOR OVERVIEW
# ═════════════════════════════════════════════════════════════════════════════
ws_ov = wb.create_sheet("COMPETITOR OVERVIEW")
ws_ov.sheet_view.showGridLines = False

col_w = [4, 22, 30, 24, 12, 32, 38, 12]
for i, w in enumerate(col_w):
    set_col(ws_ov, i+1, w)

# Title
ws_ov.merge_cells("A1:H1")
tc = ws_ov.cell(1,1, "REGENCARE.IN  —  TOP 10 COMPETITOR OVERVIEW")
tc.fill = fill(NAVY); tc.font = font(GOLD, True, 14)
tc.alignment = align("center","center",False)
set_row(ws_ov, 1, 32)

ws_ov.merge_cells("A2:H2")
sc = ws_ov.cell(2,1,"Based on Google SERP frequency analysis across 20 highest-volume keywords  ·  May 2026")
sc.fill = fill(NAVY); sc.font = font(WHITE, False, 10, italic=True)
sc.alignment = align("center","center",False)
set_row(ws_ov, 2, 20)

# Column headers
hdrs_ov = ["#","Competitor Name","Domain","Type","Threat","Locations","Primary Focus / Services","Notes"]
for c, h in enumerate(hdrs_ov, 1):
    hdr(ws_ov, 3, c, h, NAVY, WHITE, True, 10, "center")
set_row(ws_ov, 3, 30)

# Data rows
threat_colours = {"CRITICAL": CRIT, "HIGH": "C05000", "MED-HIGH": "B07700", "MEDIUM": TEAL}
threat_bgs     = {"CRITICAL": RED_L,"HIGH": AMB_L,    "MED-HIGH": AMB_L,    "MEDIUM": TEAL_L}

NOTES = {
    1: "Same doctor (Dr. Vineeth MB) — SERP cannibalization. Ranks for ortho/regen queries identical to Regencare. Immediate strategic resolution required.",
    2: "Strongest local GFC/PRP competitor. 7 clinics, 4.8★/1438 reviews. Published clinical evidence. Board-certified dermatologists. Very strong on-page SEO.",
    3: "National chain dominance. 109K+ procedures, 4.9★/1475 reviews. US-FDA approved equipment. Richest content depth. Controls hair loss + PRP SERPs in Kochi.",
    4: "International reach including UAE — directly competes in Regencare's medical tourism segment. 20+ hair services. Celebrity testimonials. 10 global locations.",
    5: "13+ branches, Dubai presence. Established 2012. Ranks for hair transplant + GFC keywords. No named doctors on pages — E-E-A-T weakness.",
    6: "Local Kochi clinic with strong GFC content. No pricing transparency. No named doctor credentials on page — exploitable E-E-A-T gap.",
    7: "Global DHI brand — #1 IMRB customer satisfaction rating. Dominates hair transplant SERPs. Not in ortho/regen segment — limited overlap with Regencare's USP.",
    8: "Only India clinic affiliated with Surecell Australia for regen medicine protocols. Ortho + stem cell focus overlaps directly with Regencare. Low online authority currently.",
    9: "Specialist GFC/hair clinic. Limited web presence and authority. Ranks for some GFC terms. Easy to outrank with stronger content.",
    10:"7 years Kerala experience. PRP + Stem Cell + Prolotherapy — direct ortho regen overlap. Low domain authority. Outranking feasible with targeted content.",
}

for i, (cid, name, domain, ctype, tier, locs, focus, threat) in enumerate(COMPETITORS):
    r = 4 + i
    bg = LIGHT if i % 2 == 0 else WHITE
    set_row(ws_ov, r, 44)
    cell(ws_ov, r, 1, str(cid), bg, NAVY, True, 11, "center")
    cell(ws_ov, r, 2, name, bg, NAVY, True, 10, "left")
    cell(ws_ov, r, 3, domain, bg, TEAL, False, 10, "left")
    cell(ws_ov, r, 4, ctype, bg, TEXT, False, 10, "left")
    # Threat column
    t_col = threat_colours.get(tier, GREY)
    t_bg  = threat_bgs.get(tier, GREY_L)
    tc2 = ws_ov.cell(r, 5, tier)
    tc2.fill = fill(t_bg); tc2.font = font(t_col, True, 9)
    tc2.alignment = align("center","center",False); tc2.border = border()
    cell(ws_ov, r, 6, locs, bg, TEXT, False, 9, "left")
    cell(ws_ov, r, 7, focus, bg, TEXT, False, 9, "left")
    cell(ws_ov, r, 8, NOTES.get(cid,""), bg, GREY, False, 9, "left")

# ═════════════════════════════════════════════════════════════════════════════
# SHEET 3 — KEYWORD RANKING MAP
# ═════════════════════════════════════════════════════════════════════════════
ws_kw = wb.create_sheet("KEYWORD RANKING MAP")
ws_kw.sheet_view.showGridLines = False

# Column widths
set_col(ws_kw, 1, 36)   # Keyword
set_col(ws_kw, 2, 9)    # Vol
set_col(ws_kw, 3, 12)   # Intent
set_col(ws_kw, 4, 16)   # Cluster
set_col(ws_kw, 5, 11)   # Regencare
for c in range(6, 16):
    set_col(ws_kw, c, 11)

ws_kw.merge_cells("A1:O1")
tc = ws_kw.cell(1,1,"KEYWORD RANKING MAP  —  Regencare vs. Top 10 Competitors")
tc.fill = fill(NAVY); tc.font = font(GOLD,True,14)
tc.alignment = align("center","center",False); set_row(ws_kw,1,32)

ws_kw.merge_cells("A2:O2")
sc = ws_kw.cell(2,1,"Positions estimated from SERP frequency analysis · Green=#1-3 · Amber=#4-7 · Red=#8-10 · 'Not Ranking'=Not on page 1")
sc.fill = fill(NAVY); sc.font = font(WHITE,False,10,italic=True)
sc.alignment = align("center","center",False); set_row(ws_kw,2,20)

# Column headers row 3
COMP_NAMES_SHORT = ["Orthogen","DermaVue","Oliva","Cutis","La Densitae","Zaaya","DHI","DH Clinic","HairWell","Epione"]
hdr(ws_kw, 3, 1, "Keyword", NAVY, WHITE, True, 9, "left")
hdr(ws_kw, 3, 2, "Est. Vol/mo", NAVY, WHITE, True, 9, "center")
hdr(ws_kw, 3, 3, "Intent", NAVY, WHITE, True, 9, "center")
hdr(ws_kw, 3, 4, "Cluster", NAVY, WHITE, True, 9, "center")
hdr(ws_kw, 3, 5, "REGENCARE", TEAL, WHITE, True, 9, "center")
for ci, cname in enumerate(COMP_NAMES_SHORT, 6):
    hdr(ws_kw, 3, ci, cname, NAVY, WHITE, True, 9, "center")
set_row(ws_kw, 3, 36)

# Data rows
for ki, (kw, vol, intent, cluster) in enumerate(KEYWORDS):
    r = 4 + ki
    bg = LIGHT if ki % 2 == 0 else WHITE
    set_row(ws_kw, r, 22)
    cell(ws_kw, r, 1, kw, bg, TEXT, False, 10, "left")
    cell(ws_kw, r, 2, f"{vol:,}", bg, NAVY, True, 10, "center")
    intent_col = GREEN if "Transact" in intent else TEAL if "Commercial" in intent else AMB
    ic = ws_kw.cell(r, 3, intent)
    ic.fill = fill(bg); ic.font = font(intent_col, False, 9)
    ic.alignment = align("center","center",False); ic.border = border()
    cell(ws_kw, r, 4, cluster, bg, TEXT, False, 9, "center")
    pos_cell(ws_kw, r, 5, REGENCARE_POS[ki])
    for ci, cid in enumerate([1,2,3,4,5,6,7,8,9,10], 6):
        pos_cell(ws_kw, r, ci, POSITIONS[cid][ki])

# Summary row — how many keywords each competitor appears in
sum_r = 4 + len(KEYWORDS)
set_row(ws_kw, sum_r, 28)
merge_hdr(ws_kw, sum_r, 1, sum_r, 4, "Keywords with Page-1 Presence (of 20)", NAVY, WHITE, 10)
regen_cnt = sum(1 for p in REGENCARE_POS if p != "NR")
pos_cell(ws_kw, sum_r, 5, str(regen_cnt))
ws_kw.cell(sum_r, 5).value = f"{regen_cnt}/20"
for ci, cid in enumerate([1,2,3,4,5,6,7,8,9,10], 6):
    cnt = sum(1 for p in POSITIONS[cid] if p != "NR")
    c2 = ws_kw.cell(sum_r, ci, f"{cnt}/20")
    c2.fill = fill(GREEN_L if cnt >= 5 else AMB_L if cnt >= 3 else RED_L)
    c2.font = font(GREEN if cnt >= 5 else AMB if cnt >= 3 else RED, True, 10)
    c2.alignment = align("center","center",False); c2.border = border()

# Freeze panes
ws_kw.freeze_panes = "E4"

# ═════════════════════════════════════════════════════════════════════════════
# SHEET 4 — ON-PAGE SEO AUDIT
# ═════════════════════════════════════════════════════════════════════════════
ws_op = wb.create_sheet("ON-PAGE SEO AUDIT")
ws_op.sheet_view.showGridLines = False

AUDIT_COLS = ["Title Tag","Single H1","Meta Desc","Schema Markup","FAQ Page","Word Count","Pricing Listed","Real Booking","Mobile-Ready","NMC Compliant"]
AUDIT_W    = [24, 22, 14, 12, 14, 14, 12, 14, 16, 16, 16]

for i, w in enumerate(AUDIT_W):
    set_col(ws_op, i+1, w)

ws_op.merge_cells(f"A1:{get_column_letter(len(AUDIT_COLS)+1)}1")
tc = ws_op.cell(1,1,"ON-PAGE SEO AUDIT  —  Regencare vs. Top 10 Competitors")
tc.fill = fill(NAVY); tc.font = font(GOLD,True,14)
tc.alignment = align("center","center",False); set_row(ws_op,1,32)

ws_op.merge_cells(f"A2:{get_column_letter(len(AUDIT_COLS)+1)}2")
sc = ws_op.cell(2,1,"Audited via direct site fetch and SERP analysis  ·  ✓=Present/Good  ~=Partial  ✗=Missing  ?=Unknown")
sc.fill = fill(NAVY); sc.font = font(WHITE,False,10,italic=True)
sc.alignment = align("center","center",False); set_row(ws_op,2,20)

# Header row
hdr(ws_op, 3, 1, "Competitor", NAVY, WHITE, True, 10, "left")
for ci, col_name in enumerate(AUDIT_COLS, 2):
    hdr(ws_op, 3, ci, col_name, NAVY, WHITE, True, 9, "center")
set_row(ws_op, 3, 36)

# Regencare row
REGENCARE_ONPAGE = ("Partial (7 H1 issue)", "No (7 H1 tags)", "Partial", "Partial", "No", 3200, "Partial", "No (form only)", "Yes", "Yes")

for r_idx, (label, data) in enumerate([("REGENCARE.IN (Client)", REGENCARE_ONPAGE)] +
    [(f"#{cid}  {name}", ONPAGE[cid]) for cid, name, *_ in COMPETITORS]):
    r = 4 + r_idx
    bg = TEAL_L if r_idx == 0 else (LIGHT if r_idx % 2 == 1 else WHITE)
    set_row(ws_op, r, 32)
    c0 = ws_op.cell(r, 1, label)
    c0.fill = fill(bg); c0.font = font(TEAL if r_idx==0 else NAVY, True, 10)
    c0.alignment = align("left","center"); c0.border = border()

    title_opt, h1, meta, schema, faq, wc, pricing, booking, mobile, nmc = data
    for ci, (val, col_idx) in enumerate(zip([title_opt, h1, meta, schema, faq, wc, pricing, booking, mobile, nmc], range(2,12))):
        if isinstance(val, int):  # word count
            wc_bg = GREEN_L if val >= 3000 else AMB_L if val >= 1500 else RED_L
            wc_fg = GREEN if val >= 3000 else AMB if val >= 1500 else RED
            cv = ws_op.cell(r, col_idx, f"{val:,} words")
            cv.fill = fill(wc_bg); cv.font = font(wc_fg, True, 10)
            cv.alignment = align("center","center",False); cv.border = border()
        else:
            tick(ws_op, r, col_idx, str(val))

# ═════════════════════════════════════════════════════════════════════════════
# SHEET 5 — AUTHORITY & TRUST
# ═════════════════════════════════════════════════════════════════════════════
ws_auth = wb.create_sheet("AUTHORITY & TRUST")
ws_auth.sheet_view.showGridLines = False

AUTH_COLS = ["Google Rating","Review Count","Years Active","Est. Domain Authority","Backlink Profile","Doctor Credentials on Page","Content Authority","E-E-A-T Score"]
AUTH_W    = [26, 14, 13, 11, 13, 16, 42, 18, 14]
for i, w in enumerate(AUTH_W):
    set_col(ws_auth, i+1, w)

ws_auth.merge_cells(f"A1:{get_column_letter(len(AUTH_COLS)+1)}1")
tc = ws_auth.cell(1,1,"AUTHORITY & TRUST SIGNALS  —  Regencare vs. Top 10 Competitors")
tc.fill = fill(NAVY); tc.font = font(GOLD,True,14)
tc.alignment = align("center","center",False); set_row(ws_auth,1,32)

ws_auth.merge_cells(f"A2:{get_column_letter(len(AUTH_COLS)+1)}2")
sc = ws_auth.cell(2,1,"E-E-A-T = Experience, Expertise, Authoritativeness, Trustworthiness  ·  DA = Domain Authority estimate  ·  Higher = Better")
sc.fill = fill(NAVY); sc.font = font(WHITE,False,10,italic=True)
sc.alignment = align("center","center",False); set_row(ws_auth,2,20)

hdr(ws_auth, 3, 1, "Competitor", NAVY, WHITE, True, 10, "left")
for ci, col_name in enumerate(AUTH_COLS, 2):
    hdr(ws_auth, 3, ci, col_name, NAVY, WHITE, True, 9, "center")
set_row(ws_auth, 3, 36)

REGEN_AUTH = (4.2, 150, "2021+", 18, "Low", "Dr. Vineeth MB — MBBS, MS Ortho (dedicated profile page)", "Medium")
EEAT_SCORES= {
    "REGENCARE.IN (Client)": 3,
    1: 3, 2: 4, 3: 5, 4: 3, 5: 2, 6: 2, 7: 4, 8: 2, 9: 1, 10: 2,
}

EEAT_LABEL = {5:"Very High", 4:"High", 3:"Medium", 2:"Low", 1:"Very Low"}

for r_idx, (label, data, eid) in enumerate(
    [("REGENCARE.IN (Client)", REGEN_AUTH, "REGENCARE.IN (Client)")] +
    [(f"#{cid}  {name}", AUTHORITY[cid], cid) for cid, name, *_ in COMPETITORS]):
    r = 4 + r_idx
    bg = TEAL_L if r_idx == 0 else (LIGHT if r_idx % 2 == 1 else WHITE)
    set_row(ws_auth, r, 36)
    c0 = ws_auth.cell(r, 1, label)
    c0.fill = fill(bg); c0.font = font(TEAL if r_idx==0 else NAVY, True, 10)
    c0.alignment = align("left","center"); c0.border = border()

    if r_idx == 0:
        rating, rev, yrs, da, bl, doc_cred, content_auth = data
    else:
        rating, rev, yrs, da, bl, doc_cred, content_auth = data

    # Rating
    if rating:
        rat_bg = GREEN_L if rating >= 4.7 else AMB_L if rating >= 4.3 else RED_L
        rat_fg = GREEN if rating >= 4.7 else AMB if rating >= 4.3 else RED
        rc = ws_auth.cell(r, 2, f"{rating} ★")
        rc.fill = fill(rat_bg); rc.font = font(rat_fg, True, 11)
        rc.alignment = align("center","center",False); rc.border = border()
    else:
        tick(ws_auth, r, 2, "Unknown")

    # Review count
    if rev:
        rv_bg = GREEN_L if rev >= 1000 else AMB_L if rev >= 300 else RED_L
        rv_fg = GREEN if rev >= 1000 else AMB if rev >= 300 else RED
        rvc = ws_auth.cell(r, 3, f"{rev:,}+")
        rvc.fill = fill(rv_bg); rvc.font = font(rv_fg, True, 10)
        rvc.alignment = align("center","center",False); rvc.border = border()
    else:
        tick(ws_auth, r, 3, "Unknown")

    cell(ws_auth, r, 4, yrs, bg, TEXT, False, 10, "center")

    # DA
    da_bg = GREEN_L if da >= 40 else AMB_L if da >= 20 else RED_L
    da_fg = GREEN if da >= 40 else AMB if da >= 20 else RED
    dac = ws_auth.cell(r, 5, str(da))
    dac.fill = fill(da_bg); dac.font = font(da_fg, True, 10)
    dac.alignment = align("center","center",False); dac.border = border()

    tick(ws_auth, r, 6, bl)
    cell(ws_auth, r, 7, doc_cred, bg, TEXT, False, 9, "left")
    tick(ws_auth, r, 8, content_auth)

    # E-E-A-T
    score = EEAT_SCORES.get(eid, 2)
    sc2 = score_cell(ws_auth, r, 9, score, EEAT_LABEL.get(score,""))
    ws_auth.cell(r, 9).font = font(
        GREEN if score>=4 else AMB if score==3 else RED, True, 10)

# ═════════════════════════════════════════════════════════════════════════════
# SHEET 6 — SERVICE COVERAGE MAP
# ═════════════════════════════════════════════════════════════════════════════
ws_svc = wb.create_sheet("SERVICE COVERAGE")
ws_svc.sheet_view.showGridLines = False

set_col(ws_svc, 1, 32)
set_col(ws_svc, 2, 14)  # Regencare
for c in range(3, 14):
    set_col(ws_svc, c, 12)

ws_svc.merge_cells(f"A1:M1")
tc = ws_svc.cell(1,1,"SERVICE COVERAGE MAP  —  Which Treatments Each Competitor Offers")
tc.fill = fill(NAVY); tc.font = font(GOLD,True,14)
tc.alignment = align("center","center",False); set_row(ws_svc,1,32)

ws_svc.merge_cells("A2:M2")
sc = ws_svc.cell(2,1,"✓ Yes = Offered and marketed  ✗ No = Not offered  ~ Partial = Mentioned but no dedicated page")
sc.fill = fill(NAVY); sc.font = font(WHITE,False,10,italic=True)
sc.alignment = align("center","center",False); set_row(ws_svc,2,20)

hdr(ws_svc, 3, 1, "Service / Treatment", NAVY, WHITE, True, 10, "left")
hdr(ws_svc, 3, 2, "REGENCARE", TEAL, WHITE, True, 9, "center")
for ci, (cid, name, *_) in enumerate(COMPETITORS, 3):
    hdr(ws_svc, 3, ci, f"#{cid} {name[:10]}", NAVY, WHITE, True, 9, "center")
set_row(ws_svc, 3, 36)

REGEN_SERVICES = ["Yes","Yes","Yes","Yes","No","Yes","No","Yes","Yes","Yes","Partial","Yes","No","No","Partial"]

for si, svc_name in enumerate(SERVICES):
    r = 4 + si
    bg = LIGHT if si % 2 == 0 else WHITE
    set_row(ws_svc, r, 24)
    c0 = ws_svc.cell(r, 1, svc_name)
    c0.fill = fill(bg); c0.font = font(NAVY, True, 10)
    c0.alignment = align("left","center"); c0.border = border()
    tick(ws_svc, r, 2, REGEN_SERVICES[si])
    for ci, cid in enumerate([1,2,3,4,5,6,7,8,9,10], 3):
        tick(ws_svc, r, ci, SVC_DATA[cid][si])

# ═════════════════════════════════════════════════════════════════════════════
# SHEET 7 — WHITE SPACE & GAPS
# ═════════════════════════════════════════════════════════════════════════════
ws_ws = wb.create_sheet("WHITE SPACE & GAPS")
ws_ws.sheet_view.showGridLines = False
ws_ws.column_dimensions["A"].width = 14
ws_ws.column_dimensions["B"].width = 38
ws_ws.column_dimensions["C"].width = 28
ws_ws.column_dimensions["D"].width = 18
ws_ws.column_dimensions["E"].width = 14
ws_ws.column_dimensions["F"].width = 52

ws_ws.merge_cells("A1:F1")
tc = ws_ws.cell(1,1,"WHITE SPACE & CONTENT GAP OPPORTUNITIES  —  What Nobody Owns Yet")
tc.fill = fill(NAVY); tc.font = font(GOLD,True,14)
tc.alignment = align("center","center",False); set_row(ws_ws,1,32)

ws_ws.merge_cells("A2:F2")
sc = ws_ws.cell(2,1,"Ranked by ease × volume opportunity. These are Regencare's fastest path to page-1 rankings with new or optimised content.")
sc.fill = fill(NAVY); sc.font = font(WHITE,False,10,italic=True)
sc.alignment = align("center","center",False); set_row(ws_ws,2,20)

ws_hdrs = ["Priority","White Space / Gap Topic","Why Nobody Owns It","Est. Opportunity","Difficulty","Recommended Regencare Action"]
for ci, h in enumerate(ws_hdrs, 1):
    hdr(ws_ws, 3, ci, h, NAVY, WHITE, True, 10, "center")
set_row(ws_ws, 3, 36)

WHITE_SPACE = [
    ("CRITICAL","Ortho regenerative medicine — full content hub","All competitors doing hair. Ortho regen is owned only by Regencare + orthogencare.com (same doc)","High Volume, Low Competition","Easy","Build treatment-specific ortho pages: knee, shoulder, spine, sports injury — one page per condition, schema-marked"),
    ("CRITICAL","Stem cell therapy Kerala — dedicated content","Only 2-3 thin pages exist. No competitor has comprehensive stem cell guide","High Volume, Low Competition","Easy","Publish 2,500-word stem cell therapy guide: what it is, conditions treated, cost, process, FAQs + FAQ schema"),
    ("CRITICAL","Cost / pricing guides for all treatments","No competitor publishes transparent pricing. Users search 'GFC therapy cost Kerala' with zero strong answers","High Vol, High Intent","Easy","Create treatment cost pages: GFC cost, PRP cost, Stem Cell cost — transparent pricing = trust + transact. keyword wins"),
    ("HIGH","UAE / NRI medical tourism Kerala — dedicated segment","No Kerala regen clinic has an NRI-facing landing page. 18 keywords, near-zero competition","Medium Vol, Zero Comp.","Easy","Build /nri-medical-tourism page targeting 'regenerative medicine India for NRI', 'Kerala ortho clinic for UAE residents'"),
    ("HIGH","Dr. Vineeth MB — authority content hub","No competitor (incl. orthogencare) has a comprehensive doctor profile with original content, papers, and Q&A","Medium Vol, Low Comp.","Easy","Dr. Vineeth MB hub: credentials, media, patient Q&A, research summaries — this becomes Regencare's primary E-E-A-T signal"),
    ("HIGH","Sports injury treatment Kerala — non-surgical","Ayurveda dominates. No modern regen clinic has strong sports injury content","Medium Vol, Low Comp.","Easy","Build /sports-injury page: ACL, PCL, meniscal tears, tennis elbow — PRP/stem cell as alternative to surgery"),
    ("HIGH","GFC therapy FAQ and how-it-works content","All GFC pages are thin. DermaVue has best content but no FAQ schema. No competitor answers: sessions, pain, downtime","High Vol, High Intent","Easy","Write comprehensive 3,000-word GFC guide with FAQ schema: 15+ questions answered — targets PAA boxes + featured snippets"),
    ("HIGH","Hair regrowth for men — condition-specific Kerala","Generic hair pages dominate. No clinic has a condition-specific 'hair regrowth for men Kerala' hub","Medium Vol, Low Comp.","Medium","Build /conditions/hair-loss page: causes, options, cost, Regencare's approach — targets informational + commercial intent"),
    ("MEDIUM","Regenerative medicine research articles","Nobody publishes clinical research or outcome data in Kerala context — major E-E-A-T signal gap","Low Vol, Very Low Comp.","Medium","Dr. Vineeth MB publishes 4-6 research summaries/year on regencare.in — cites clinical studies, adds original commentary"),
    ("MEDIUM","Calicut / Kozhikode — dedicated treatment pages","All competitors focus Kochi. Calicut has lower competition for same keywords + GBP local pack opportunity","Medium Vol, Low Comp.","Easy","Build Calicut-specific treatment pages: /gfc-therapy/calicut, /prp-treatment/calicut — same content structure as Kochi pages"),
    ("MEDIUM","Chennai regenerative medicine — market entry","No competitor in Chennai has strong regen medicine content. New market, low SERP competition","Medium Vol, Low Comp.","Easy","Build Chennai-specific treatment and branch pages before competitors notice the gap"),
    ("MEDIUM","Platelet therapy vs. PRP vs. GFC — comparison content","No comparison content exists. Users search 'PRP vs GFC' with no clear authoritative answer on SERPs","Medium Vol, Low Comp.","Easy","Write comparison page: GFC vs PRP vs Stem Cell — defines Regencare as the expert, drives commercial-intent traffic"),
    ("MEDIUM","Non-surgical spine and back pain Kerala","Physiotherapy and Ayurveda own this space. Modern regen approach unclaimed","Medium Vol, Low Comp.","Medium","Build /conditions/spine-back-pain — PRP for disc herniation, prolotherapy for chronic back pain"),
    ("LOW","Google AI Overview citations for regen medicine","No Kerala clinic is being cited by ChatGPT/Gemini for regen queries — citation-bait content needed","N/A","Hard","Publish original clinical data, doctor FAQs, and structured answer content — these train AI models to cite Regencare"),
    ("LOW","Video content — procedure walkthroughs","No Kerala regen clinic has strong YouTube presence for their treatment keywords","Video SERP gap","Medium","Film GFC, PRP procedure walkthroughs with Dr. Vineeth MB — YouTube SEO + embed on treatment pages"),
]

priority_cols = {"CRITICAL": (RED_L, RED), "HIGH": (AMB_L, AMB), "MEDIUM": (TEAL_L, TEAL), "LOW": (GREY_L, GREY)}

for wi, row_data in enumerate(WHITE_SPACE):
    r = 4 + wi
    bg = LIGHT if wi % 2 == 0 else WHITE
    set_row(ws_ws, r, 44)
    priority, topic, why, opp, diff, action = row_data
    p_bg, p_fg = priority_cols.get(priority, (WHITE, TEXT))
    pc = ws_ws.cell(r, 1, priority)
    pc.fill = fill(p_bg); pc.font = font(p_fg, True, 9)
    pc.alignment = align("center","center",False); pc.border = border()
    cell(ws_ws, r, 2, topic, bg, NAVY, True, 10, "left")
    cell(ws_ws, r, 3, why, bg, TEXT, False, 9, "left")
    # Opportunity
    opp_bg = GREEN_L if "High" in opp and "Zero" not in opp else TEAL_L if "Medium" in opp else AMB_L
    oc = ws_ws.cell(r, 4, opp)
    oc.fill = fill(opp_bg); oc.font = font(TEAL, False, 9)
    oc.alignment = align("center","center",True); oc.border = border()
    # Difficulty
    diff_bg = GREEN_L if diff == "Easy" else AMB_L if diff == "Medium" else RED_L
    diff_fg = GREEN if diff == "Easy" else AMB if diff == "Medium" else RED
    dc2 = ws_ws.cell(r, 5, diff)
    dc2.fill = fill(diff_bg); dc2.font = font(diff_fg, True, 9)
    dc2.alignment = align("center","center",False); dc2.border = border()
    cell(ws_ws, r, 6, action, bg, TEXT, False, 9, "left")

# ═════════════════════════════════════════════════════════════════════════════
# SHEET 8 — ACTION PRIORITIES
# ═════════════════════════════════════════════════════════════════════════════
ws_act = wb.create_sheet("ACTION PRIORITIES")
ws_act.sheet_view.showGridLines = False
ws_act.column_dimensions["A"].width = 6
ws_act.column_dimensions["B"].width = 14
ws_act.column_dimensions["C"].width = 38
ws_act.column_dimensions["D"].width = 42
ws_act.column_dimensions["E"].width = 22
ws_act.column_dimensions["F"].width = 14
ws_act.column_dimensions["G"].width = 40

ws_act.merge_cells("A1:G1")
tc = ws_act.cell(1,1,"COMPETITIVE STUDY — ACTION PRIORITIES FOR REGENCARE")
tc.fill = fill(NAVY); tc.font = font(GOLD,True,14)
tc.alignment = align("center","center",False); set_row(ws_act,1,32)

ws_act.merge_cells("A2:G2")
sc = ws_act.cell(2,1,"Actions ranked by impact × urgency. Derived directly from competitive intelligence findings. GTM Step references match the GTM Master List document.")
sc.fill = fill(NAVY); sc.font = font(WHITE,False,10,italic=True)
sc.alignment = align("center","center",False); set_row(ws_act,2,20)

act_hdrs = ["#","Priority","Action","Competitive Insight Behind It","Owner","GTM Step","Expected Outcome"]
for ci, h in enumerate(act_hdrs, 1):
    hdr(ws_act, 3, ci, h, NAVY, WHITE, True, 10, "center")
set_row(ws_act, 3, 36)

ACTIONS = [
    (1,"CRITICAL","Resolve orthogencare.com brand cannibalization",
     "Same doctor (Dr. Vineeth MB) on competing domain appears alongside regencare.in for ortho regen queries. Every shared SERP splits authority and clicks.",
     "Strategy + Dev","Step 2A","Stop losing 30-50% of ortho regen SERP traffic to the same doctor's other domain"),
    (2,"CRITICAL","Fix 7 H1 tags on homepage",
     "DermaVue, Oliva, and Cutis all have clean, single H1. Regencare's 7 H1s confuse crawlers and dilute primary keyword signal.",
     "Dev","Step 2A","Correct keyword signal sent to Google for primary homepage query cluster"),
    (3,"CRITICAL","Restore Blog (404) and Resources (404) pages",
     "All top competitors — DermaVue (blog), Oliva (blog), Cutis (blog) — use blog content to rank for informational keywords. Regencare's blog is entirely absent.",
     "Dev + Content","Step 2A","Restore the primary content authority surface — prerequisite for all content production"),
    (4,"CRITICAL","Replace contact form with real booking system (Cal.com)",
     "DHI, Oliva, DermaVue all have real appointment booking. Regencare's 'Book' button leads to a contact form — zero conversion tracking.",
     "Dev","Step 3A","Direct patient bookings tracked and confirmed — eliminates the biggest conversion leakage point"),
    (5,"CRITICAL","Fix Chennai NAP inconsistency (email domain mismatch)",
     "NAP consistency is a local SEO ranking factor. Competitors with clean NAP dominate local packs. Regencare's Chennai listing conflicts with Google Business Profile.",
     "Marketing","Step 2A","Chennai GBP and local pack eligibility restored"),
    (6,"HIGH","Publish transparent pricing pages for GFC, PRP, Stem Cell",
     "No top-10 competitor lists pricing online. 'GFC therapy cost Kerala' (170 vol/mo) has no strong answer — instant featured snippet opportunity with a dedicated pricing page.",
     "Content","Step 3B","Rank for cost/pricing keywords with zero competition. Highest-converting traffic type."),
    (7,"HIGH","Build comprehensive GFC therapy content hub (3,000+ words + FAQ schema)",
     "Regencare ranks #9. DermaVue ranks #1 with 4,900-word page. Content depth is the gap — not domain authority. DermaVue has no FAQ schema.",
     "Content","Step 3B","Move from #9 to top 3 for GFC Kochi. Capture 5+ People Also Ask boxes with FAQ schema"),
    (8,"HIGH","Build UAE/NRI medical tourism landing page",
     "Cutis and DHI have UAE presence. None has a Kerala-based regen clinic page specifically targeting NRI searches. 18 keywords, near-zero competition.",
     "Content","Step 3A","Capture UAE segment traffic — highest LTV patient group with no current competitor"),
    (9,"HIGH","Optimise Dr. Vineeth MB profile as E-E-A-T hub",
     "La Densitae and Zaaya rank without named doctors. DermaVue's named MD credentials are a clear ranking factor. Dr. Vineeth MB's credentials are Regencare's biggest competitive moat.",
     "Content + SEO","Step 3A","Strengthen all-site E-E-A-T signal — improves ranking for every page on the domain"),
    (10,"HIGH","Implement schema markup: MedicalClinic, MedicalProcedure, FAQPage per page",
     "DermaVue has 4 schema types and outranks Regencare on GFC. Orthogencare has zero schema — Regencare should implement comprehensively to separate from both.",
     "Dev","Step 3A","SERP feature eligibility: rich results, People Also Ask, featured snippets, local pack enhancement"),
    (11,"HIGH","Build Calicut-specific treatment pages",
     "All competitors focus on Kochi. Calicut has lower keyword competition for the same queries — lower barrier to page 1 rankings.",
     "Content","Step 3B","Page-1 rankings for Calicut keywords with less content investment than equivalent Kochi pages"),
    (12,"HIGH","Optimise Google Business Profile for all 3 branches",
     "DermaVue (7 clinics, strong GBP) and Oliva dominate local packs. Regencare GBPs are incomplete — NAP issues, missing service listings, low review volume.",
     "Marketing","Step 4A","Local pack presence for Kochi, Calicut, Chennai — drives walk-in and call conversions"),
    (13,"MEDIUM","Publish stem cell therapy comprehensive guide",
     "Only 2-3 thin pages exist on this topic. Regencare and orthogencare rank but neither has a comprehensive guide. Lowest-competition, high-volume opportunity.",
     "Content","Step 3B","Top-3 ranking for 'stem cell therapy Kerala' with zero competitor currently occupying that spot with quality content"),
    (14,"MEDIUM","Launch sports injury treatment content (PRP/regen — non-surgical)",
     "Physiotherapy and Ayurveda own this space. No modern regen clinic has claimed it.",
     "Content","Step 3B","Capture untapped sports injury segment — high commercial intent, no current regen competitor"),
    (15,"MEDIUM","Film and publish GFC and PRP procedure video content",
     "Video carousels appear on GFC and hair loss SERPs. No Kerala regen clinic has strong YouTube presence. DermaVue has video on site. Competitors are absent from video SERPs.",
     "Content + Video","Step 4B","YouTube rankings + video carousel positions on Google for priority keywords"),
    (16,"LOW","Begin AI citation strategy (GEO — ChatGPT, Gemini, Perplexity)",
     "No Kerala clinic is cited by AI engines for regen queries. Publishing original clinical data and structured FAQ content trains AI models to cite Regencare.",
     "Content","Step 5C","Regencare cited in AI Overview, ChatGPT, and Perplexity for regenerative medicine Kerala queries"),
]

for r_idx, row_data in enumerate(ACTIONS):
    r = 4 + r_idx
    bg = LIGHT if r_idx % 2 == 0 else WHITE
    set_row(ws_act, r, 50)
    num, priority, action, insight, owner, gtm, outcome = row_data
    cell(ws_act, r, 1, str(num), bg, NAVY, True, 11, "center")
    p_bg, p_fg = priority_cols.get(priority, (WHITE, TEXT))
    pc = ws_act.cell(r, 2, priority)
    pc.fill = fill(p_bg); pc.font = font(p_fg, True, 9)
    pc.alignment = align("center","center",False); pc.border = border()
    cell(ws_act, r, 3, action, bg, NAVY, True, 10, "left")
    cell(ws_act, r, 4, insight, bg, TEXT, False, 9, "left")
    cell(ws_act, r, 5, owner, bg, TEAL, True, 9, "center")
    cell(ws_act, r, 6, gtm, bg, NAVY, True, 9, "center")
    cell(ws_act, r, 7, outcome, bg, GREEN, False, 9, "left")

# ── FOOTER on action sheet ──
foot_r = 4 + len(ACTIONS) + 1
ws_act.merge_cells(f"A{foot_r}:G{foot_r}")
fc = ws_act.cell(foot_r, 1, "Prepared by NT Global Digital  ·  Client: Regencare.in  ·  Competitive Study v1.0  ·  May 2026  ·  Confidential")
fc.fill = fill(NAVY); fc.font = font(GOLD, False, 9, italic=True)
fc.alignment = align("center","center",False)

# ═════════════════════════════════════════════════════════════════════════════
# SAVE
# ═════════════════════════════════════════════════════════════════════════════
out = r"c:\project\AI RESEARCH INTELLIGENCE SYSTEM\Regencare_Competitive_Study.xlsx"
wb.save(out)
print(f"Saved: {out}")
