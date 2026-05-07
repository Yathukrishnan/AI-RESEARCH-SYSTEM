import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import datetime

wb = openpyxl.Workbook()

# ── Palette ─────────────────────────────────────────────────────
NAVY        = "1A3C5E"
TEAL        = "1F6F8B"
TEAL_LIGHT  = "D6EAF8"
GREEN       = "1E8449"
GREEN_LT    = "D5F5E3"
RED         = "C0392B"
RED_LT      = "FADBD8"
AMBER       = "B7770D"
AMBER_LT    = "FDEBD0"
GREY        = "F4F6F7"
WHITE       = "FFFFFF"
DARK_GREY   = "717D7E"
GOLD        = "D4AC0D"

def fill(c): return PatternFill("solid", fgColor=c)
def fnt(c=WHITE, sz=10, bold=False, italic=False):
    return Font(name="Calibri", color=c, size=sz, bold=bold, italic=italic)
def bdr():
    s = Side(style="thin", color="D5D8DC")
    return Border(left=s, right=s, top=s, bottom=s)
def aln(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def pos_fill(p):
    p = str(p)
    if p in ("#1","#2","#3"):         return fill(GREEN_LT)
    if any(p == f"#{i}" for i in range(4,11)): return fill(AMBER_LT)
    return fill(RED_LT)
def pos_fnt(p):
    p = str(p)
    if p in ("#1","#2","#3"):         return fnt(GREEN,  10, bold=True)
    if any(p == f"#{i}" for i in range(4,11)): return fnt(AMBER, 10, bold=True)
    return fnt(RED, 10, bold=True)

COLS = ["Keyword","Current Position","Regencare URL Ranking",
        "Search Intent","Top Competitor","Est. Monthly Searches",
        "Priority","Action Required"]
COL_W = [44, 14, 48, 18, 32, 20, 10, 56]

def hdr_row(ws, r, vals):
    for c, v in enumerate(vals, 1):
        cl = ws.cell(r, c, v)
        cl.fill = fill(NAVY); cl.font = fnt(WHITE,10,bold=True)
        cl.alignment = aln("center"); cl.border = bdr()
    ws.row_dimensions[r].height = 26

def data_row(ws, r, vals):
    for c, v in enumerate(vals, 1):
        cl = ws.cell(r, c, v)
        cl.border = bdr()
        if c == 2:
            cl.fill = pos_fill(v); cl.font = pos_fnt(v)
            cl.alignment = aln("center", wrap=False)
        elif c == 7:
            pc = {
                "Critical": (RED_LT,   RED,   True),
                "High":     (AMBER_LT, AMBER, True),
                "Medium":   (TEAL_LIGHT,TEAL, False),
                "Low":      (GREY,   DARK_GREY, False),
            }.get(str(v), (WHITE, "000000", False))
            cl.fill = fill(pc[0]); cl.font = fnt(pc[1],10,bold=pc[2])
            cl.alignment = aln("center", wrap=False)
        else:
            cl.fill = fill(GREY) if r % 2 == 0 else fill(WHITE)
            cl.font = fnt("222222", 10)
            cl.alignment = aln()
    ws.row_dimensions[r].height = 22

def set_col_widths(ws):
    for i, w in enumerate(COL_W, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def summary_bar(ws, r, total, r1, r2, nt):
    ws.merge_cells(f"A{r}:H{r}")
    cl = ws[f"A{r}"]
    cl.value = (f"  TOTAL: {total} keywords     "
                f"Ranking #1–#3: {r1}     "
                f"Ranking #4–#10: {r2}     "
                f"Not in Top 10 (Gap): {nt}")
    cl.fill = fill(NAVY); cl.font = fnt(WHITE,10,bold=True)
    cl.alignment = aln("left"); ws.row_dimensions[r].height = 24

def build_sheet(wb, title, tab_color, data):
    ws = wb.create_sheet(title)
    ws.sheet_properties.tabColor = tab_color
    ws.freeze_panes = "A4"

    # Title banner
    ws.merge_cells("A1:H1")
    cl = ws["A1"]
    cl.value = f"REGENCARE.IN  ·  KEYWORD BASELINE REPORT  ·  {title.upper()}  ·  April 2026"
    cl.fill = fill(NAVY); cl.font = fnt(WHITE, 12, bold=True)
    cl.alignment = aln("center"); ws.row_dimensions[1].height = 32

    # Legend
    ws.merge_cells("A2:H2")
    cl = ws["A2"]
    cl.value = ("   GREEN = Ranking #1–#3  (protect & optimise)     "
                "AMBER = Ranking #4–#10  (improvable)     "
                "RED = Not in Top 10  (gap / opportunity)")
    cl.fill = fill(TEAL_LIGHT); cl.font = fnt(NAVY, 9, italic=True)
    cl.alignment = aln("center"); ws.row_dimensions[2].height = 18

    hdr_row(ws, 3, COLS)
    for i, row in enumerate(data, 4):
        data_row(ws, i, row)

    r1 = sum(1 for r in data if str(r[1]) in ("#1","#2","#3"))
    r2 = sum(1 for r in data if any(str(r[1])==f"#{i}" for i in range(4,11)))
    nt = sum(1 for r in data if "Not" in str(r[1]) or "Listed" in str(r[1]))
    summary_bar(ws, len(data)+5, len(data), r1, r2, nt)

    set_col_widths(ws)
    return ws

# ═══════════════════════════════════════════════════════════════
# DATA
# ═══════════════════════════════════════════════════════════════

KOCHI = [
    ("regenerative medicine Kochi",              "#1",  "regencare.in",                                        "Brand",        "orthogencare.com",              "1,000–2,000",  "High",     "Add MedicalClinic schema — lock in position"),
    ("PRP therapy Kochi Kerala",                 "#1",  "regencare.in",                                        "Transactional","orthogencare.com",              "1,000–2,000",  "High",     "Add FAQPage schema to secure featured snippet"),
    ("hair loss regenerative medicine Kochi",    "#1",  "regencare.in",                                        "Transactional","DermaVue",                      "500–1,000",    "High",     "Build dedicated hair loss landing page"),
    ("erectile dysfunction treatment Kochi",     "#1",  "regencare.in/erectile-dysfunction-regenerative-medicine","Transactional","Silverline Hospital",          "500–1,000",    "Medium",   "Sensitive page — needs DPDP compliance review"),
    ("osteoarthritis treatment without surgery Kerala","#1","regencare.in/ortho/osteochondritis-treatment-kerala","Transactional","Ayurveda clinics",            "1,000–2,000",  "High",     "Add MedicalProcedure schema immediately"),
    ("avascular necrosis treatment Kerala",      "#1",  "regencare.in/ortho/avascular-necrosis-treatment",    "Transactional","Chaitanya Stem Cell",           "200–500",      "Medium",   "Niche high-intent — add schema"),
    ("can I avoid knee surgery with PRP Kerala", "#1",  "regencare.in/ortho/knee-pain-treatment-without-surgery","Question",   "orthogencare.com",              "1,000–2,000",  "High",     "Perfect intent match — add FAQ schema immediately"),
    ("regenerative medicine near me Kerala",     "#1",  "regencare.in",                                        "Navigational", "orthogencare.com, Epione",      "2,000–5,000",  "High",     "'Near me' high intent — optimise all 3 GBP profiles"),
    ("knee pain treatment without surgery Kochi","#2",  "regencare.in/ortho/knee-pain-treatment-without-surgery","Transactional","Physiotherapy blog",           "2,000–5,000",  "High",     "Strong opportunity — schema + content to claim #1"),
    ("stem cell therapy Kochi Kerala",           "#2",  "regencare.in/stem-cell-therapy",                     "Transactional","orthogencare.com (#1)",         "500–1,000",    "Critical", "Brand conflict — orthogencare.com #1, resolve urgently"),
    ("tendonitis treatment Kerala regenerative", "#3",  "regencare.in/ortho/chronic-tendonitis-treatment",    "Informational","orthogencare.com",              "200–500",      "Medium",   "Add FAQ section to improve position"),
    ("shin splints treatment Kochi Kerala",      "#3",  "regencare.in/ortho/shin-splints-treatment",          "Transactional","orthogencare.com (#2)",         "200–500",      "Medium",   "Brand conflict — orthogencare.com above"),
    ("non surgical treatment near me Kochi",     "#3",  "regencare.in/about-us",                              "Transactional","Almeka, Skinessence",           "1,000–2,000",  "High",     "About Us ranking — build dedicated non-surgical hub page"),
    ("GFC therapy Kochi",                        "#9",  "regencare.in/gfc-therapy",                           "Transactional","Almeka, Hair Wellness, DermaVue","1,000–2,000",  "High",     "5 hair clinics above — rewrite page for ortho + hair GFC"),
    ("PRP treatment Kochi",                      "Not in Top 10","—",                                         "Transactional","DermaVue, Oliva, Hair Tree",    "5,000–10,000", "Critical", "Highest volume keyword — hair clinics dominate, create dedicated PRP page"),
    ("hair fall doctor Kochi",                   "Not in Top 10","—",                                         "Transactional","DermaVue, Oliva, Twacha",       "2,000–5,000",  "High",     "High patient volume — build Dr. Aswathi profile page"),
    ("knee pain doctor Kochi",                   "Not in Top 10","—",                                         "Transactional","Lybrate, Amrita, Lourdes",      "2,000–5,000",  "High",     "Very high intent — create doctor profile with ortho keywords"),
    ("joint pain treatment Kochi",               "Not in Top 10","—",                                         "Transactional","orthogencare.com, Ayurveda",    "2,000–5,000",  "High",     "Build joint pain landing page with LocalBusiness schema"),
    ("back pain treatment Kochi",                "Not in Top 10","—",                                         "Transactional","Kumar Centre, Ayurveda",        "2,000–5,000",  "High",     "No page exists — create spine/back pain condition page"),
    ("shoulder pain treatment Kochi",            "Not in Top 10","—",                                         "Transactional","orthogencare.com, Ayurveda",    "1,000–2,000",  "High",     "orthogencare.com has dedicated page — Regencare does not"),
    ("sports injury doctor Kochi",               "Not in Top 10","—",                                         "Transactional","MIOC Ortho, Aster, VPS",        "2,000–5,000",  "High",     "Hospital brands dominate — need sports medicine content hub"),
    ("hair thinning treatment Kochi",            "Not in Top 10","—",                                         "Transactional","DermaVue, Oliva, Zaaya",        "1,000–2,000",  "High",     "High volume — hair clinics fully dominate, need trichology page"),
    ("knee replacement alternative Kochi",       "Not in Top 10","—",                                         "Transactional","Kinder Hospital, VPS",          "1,000–2,000",  "High",     "Core Regencare audience — patients avoiding surgery"),
    ("pain relief clinic Kochi",                 "Not in Top 10","—",                                         "Transactional","Amrita, Kumar Centre, Aster",   "2,000–5,000",  "High",     "Hospital brands dominate — need pain management page"),
    ("skin doctor Kochi dermatologist",          "Not in Top 10","—",                                         "Navigational", "Practo, Aster, Twacha",         "5,000–10,000", "Medium",   "Very high volume — Dr. Aswathi profile page with schema needed"),
    ("acne scar treatment Kochi",                "Not in Top 10","—",                                         "Transactional","Skinessence, Oliva, DermaVue",  "2,000–5,000",  "Medium",   "High volume — dedicated acne scar treatment page needed"),
    ("plantar fasciitis treatment Kochi",        "Not in Top 10","—",                                         "Transactional","orthogencare.com",              "500–1,000",    "Medium",   "orthogencare.com has dedicated page — Regencare must match"),
    ("orthopedic doctor Kochi without surgery",  "Not in Top 10","—",                                         "Transactional","MIOC Ortho, Amrita",            "1,000–2,000",  "High",     "Perfect Regencare audience — non-surgical ortho page needed"),
    ("diabetic foot ulcer treatment Kochi",      "Not in Top 10","—",                                         "Transactional","Bethaniya Clinic, Medfin",      "500–1,000",    "Medium",   "Service exists — page not optimised for this keyword"),
    ("rotator cuff injury treatment Kochi",      "Not in Top 10","—",                                         "Transactional","International sites dominate",  "500–1,000",    "Medium",   "Regencare treats this — no dedicated page exists"),
    ("hair fall PRP Kochi",                      "Not in Top 10","—",                                         "Transactional","Oliva, Pristyn, DermaVue",      "2,000–5,000",  "High",     "High intent — regencare.in/gfc-therapy not targeting this"),
]

CALICUT = [
    ("regenerative medicine Calicut",            "#1",  "regencare.in",                                        "Brand",        "D Lapp Hair Clinic",            "500–1,000",    "High",     "Ranking #1 — optimise GBP Calicut branch"),
    ("regenerative medicine Kozhikode",          "#1",  "regencare.in",                                        "Brand",        "D Lapp Hair Clinic",            "500–1,000",    "High",     "Ranking #1 — add LocalBusiness schema for Calicut"),
    ("PRP treatment Calicut",                    "#2",  "regencare.in",                                        "Transactional","Face360, BHC, DHI India",       "1,000–2,000",  "High",     "Hair clinics above — create dedicated Calicut PRP page"),
    ("PRP treatment Kozhikode",                  "#2",  "regencare.in",                                        "Transactional","Face360, Hair Tree",            "500–1,000",    "High",     "Create dedicated Kozhikode branch treatment page"),
    ("knee pain treatment Calicut",              "Not in Top 10","—",                                         "Transactional","Meitra Hospital, GMC, Lybrate", "2,000–5,000",  "High",     "Hospital brands dominate — need Calicut knee condition page"),
    ("knee pain doctor Calicut Kozhikode",       "Not in Top 10","—",                                         "Transactional","Meitra Hospital, Lybrate",      "1,000–2,000",  "High",     "High intent — Dr. Vineeth Calicut profile page needed"),
    ("hair loss treatment Calicut",              "Not in Top 10","—",                                         "Transactional","DHI, La Densitae, Charmam",     "1,000–2,000",  "High",     "Hair transplant clinics dominate — build hair treatment Calicut page"),
    ("sports injury treatment Calicut",          "Not in Top 10","—",                                         "Transactional","Meitra Hospital, GMC",          "500–1,000",    "Medium",   "Hospital brands dominate — build sports medicine Calicut content"),
    ("hair fall doctor Calicut Kozhikode",       "Not in Top 10","—",                                         "Transactional","Practo, Lybrate, Charmam",      "500–1,000",    "High",     "No Regencare presence — add dermatologist Calicut profile"),
    ("GFC therapy Calicut",                      "Not in Top 10","—",                                         "Transactional","Hair clinics, DHI",             "500–1,000",    "High",     "Regencare offers GFC in Calicut — no dedicated Calicut GFC page"),
    ("joint pain treatment Calicut",             "Not in Top 10","—",                                         "Transactional","Meitra Hospital, Lybrate",      "500–1,000",    "High",     "Create Calicut joint pain landing page with local schema"),
    ("stem cell therapy Calicut",                "Not in Top 10","—",                                         "Transactional","Not many competitors",          "200–500",      "Medium",   "Low competition — opportunity to rank with dedicated page"),
    ("osteoarthritis treatment Calicut",         "Not in Top 10","—",                                         "Transactional","Ayurveda clinics",              "500–1,000",    "High",     "Ayurveda dominates — Regencare's non-surgical approach differentiates"),
    ("back pain treatment Kozhikode",            "Not in Top 10","—",                                         "Transactional","Maana Health, Ayurveda",        "1,000–2,000",  "High",     "High volume — spine treatment page for Calicut needed"),
    ("skin clinic Calicut Kozhikode",            "Not in Top 10","—",                                         "Navigational", "Criniere Dermis, DHI",          "1,000–2,000",  "Medium",   "Dermatology presence — Dr. Aswathi Calicut page needed"),
    ("PRP cost Calicut Kozhikode",               "Not in Top 10","—",                                         "Transactional","Various hair clinics",          "500–1,000",    "High",     "Cost comparison page for Calicut branch needed"),
]

CHENNAI = [
    ("PRP treatment Chennai regenerative",       "#3",  "regencare.in",                                        "Transactional","Regen Ortho Care (#1)",         "2,000–5,000",  "High",     "Dedicated Chennai PRP page can push to #1"),
    ("regenerative medicine Chennai",            "#4",  "regencare.in",                                        "Brand",        "Regen Ortho Care, IIRM",        "1,000–2,000",  "High",     "Chennai branch page with full local schema needed"),
    ("stem cell therapy Chennai",                "Not in Top 10","—",                                         "Transactional","Tosh Hospital, Apollo Spectra",  "2,000–5,000",  "High",     "Hospital brands dominate — build Chennai stem cell page"),
    ("GFC therapy Chennai",                      "Not in Top 10","—",                                         "Transactional","VK Allure, Mahi Clinic",        "2,000–5,000",  "High",     "Hair clinics dominate — Regencare not appearing at all"),
    ("joint pain treatment Chennai non surgical","Not in Top 10","—",                                         "Transactional","Epione Pain, Synapse Clinic",   "2,000–5,000",  "High",     "Epione dominates — Chennai ortho condition page needed"),
    ("knee pain treatment Chennai",              "Not in Top 10","—",                                         "Transactional","Epione, Bharat Ortho, Rathi",   "5,000–10,000", "Critical", "Very high volume — Chennai knee pain page is a major gap"),
    ("hair loss treatment Chennai",              "Not in Top 10","—",                                         "Transactional","Multiple hair clinics",          "5,000–10,000", "High",     "Very high volume — need Chennai hair treatment landing page"),
    ("PRP hair treatment Chennai",               "Not in Top 10","—",                                         "Transactional","Mahi Clinic, Ram Skin",         "2,000–5,000",  "High",     "High intent — no Regencare Chennai hair page exists"),
    ("sports injury treatment Chennai",          "Not in Top 10","—",                                         "Transactional","Practo, Hospital brands",       "2,000–5,000",  "Medium",   "Hospital brands dominate — need sports medicine Chennai page"),
    ("skin doctor Chennai dermatologist",        "Not in Top 10","—",                                         "Navigational", "Practo, Hospitals",             "10,000+",      "Medium",   "Extremely high volume — Dr. Aswathi Chennai profile needed"),
    ("osteoarthritis treatment Chennai",         "Not in Top 10","—",                                         "Transactional","Epione, Bharat Ortho",          "2,000–5,000",  "High",     "Build Chennai condition pages — large elderly population"),
    ("non surgical orthopedic Chennai",          "Not in Top 10","—",                                         "Transactional","Epione Pain and Spine",         "1,000–2,000",  "High",     "Epione dominates — Regencare has no Chennai ortho page"),
    ("regenerative medicine doctor Chennai",     "Not in Top 10","—",                                         "Navigational", "Regen Ortho Care",              "500–1,000",    "High",     "Doctor profile page for Chennai branch needed with Person schema"),
    ("plantar fasciitis treatment Chennai",      "Not in Top 10","—",                                         "Transactional","Practo, Hospital brands",       "500–1,000",    "Medium",   "Build condition pages for Chennai branch"),
    ("avascular necrosis treatment Chennai",     "Not in Top 10","—",                                         "Transactional","Practo, Hospitals",             "200–500",      "Medium",   "Extend AVN page to target Chennai"),
    ("PRP cost Chennai price",                   "Not in Top 10","—",                                         "Transactional","SBJ Ortho, Practo",             "2,000–5,000",  "High",     "High intent — patients compare costs before booking in Chennai"),
]

KERALA_REGIONAL = [
    ("regenerative medicine clinic Kerala",      "#1",  "regencare.in",                                        "Informational","DH Clinic, Epione",             "2,000–5,000",  "High",     "Core brand keyword — add MedicalClinic schema"),
    ("PRP therapy Kerala",                       "#1",  "regencare.in",                                        "Transactional","orthogencare.com",              "2,000–5,000",  "High",     "Add FAQPage schema to protect position"),
    ("osteoarthritis treatment Kerala",          "#1",  "regencare.in/ortho/osteochondritis-treatment-kerala","Transactional","Ayurveda clinics",              "2,000–5,000",  "High",     "Ranking #1 — add schema and expand FAQ section"),
    ("avascular necrosis treatment Kerala",      "#1",  "regencare.in/ortho/avascular-necrosis-treatment",    "Transactional","Chaitanya Stem Cell",           "500–1,000",    "Medium",   "Niche high intent — protect with schema"),
    ("stem cell therapy Kerala",                 "#2",  "regencare.in/stem-cell-therapy",                     "Transactional","orthogencare.com (#1)",         "1,000–2,000",  "Critical", "Brand conflict — orthogencare.com #1, resolve immediately"),
    ("non surgical orthopedic treatment Kerala", "#2",  "regencare.in/about-us",                              "Transactional","Physiotherapy clinics",         "1,000–2,000",  "High",     "About Us page ranking — build dedicated non-surgical hub page"),
    ("knee pain without surgery Kerala",         "#2",  "regencare.in/ortho/knee-pain-treatment-without-surgery","Transactional","Physiotherapy blog",          "2,000–5,000",  "High",     "Strong opportunity to reach #1 with schema"),
    ("GFC therapy Kerala",                       "#3",  "regencare.in/gfc-therapy",                           "Transactional","Hair clinics",                  "1,000–2,000",  "High",     "Hair clinics above — expand GFC page to cover ortho + hair"),
    ("rheumatoid arthritis treatment Kerala",    "Not in Top 10","—",                                         "Transactional","Ayurveda clinics dominate",     "2,000–5,000",  "High",     "Ayurveda owns this — Regencare regenerative angle underserved"),
    ("diabetic foot ulcer treatment Kerala",     "Not in Top 10","—",                                         "Transactional","Bethaniya, Sanjeevanam",        "1,000–2,000",  "Medium",   "Service exists — dedicated page with schema needed"),
    ("sports injury treatment Kerala",           "Not in Top 10","—",                                         "Transactional","Aster, Hospital brands",        "2,000–5,000",  "High",     "Hospital brands dominate — sports medicine content hub needed"),
    ("hair fall treatment Kerala PRP",           "Not in Top 10","—",                                         "Transactional","DermaVue, Oliva",               "5,000–10,000", "Critical", "Very high volume — Regencare absent from this query"),
    ("joint pain doctor Kerala",                 "Not in Top 10","—",                                         "Transactional","Hospital brands",               "2,000–5,000",  "High",     "High intent — build Kerala joint pain hub with doctor profiles"),
    ("back pain specialist Kerala",              "Not in Top 10","—",                                         "Transactional","Kumar Centre, Maana",           "2,000–5,000",  "High",     "High volume — spine treatment content needed"),
    ("skin specialist Kerala dermatologist",     "Not in Top 10","—",                                         "Navigational", "Aster, Hospitals, Practo",      "5,000–10,000", "Medium",   "Very high volume — Dr. Aswathi Kerala profile page needed"),
    ("plantar fasciitis treatment Kerala",       "Not in Top 10","—",                                         "Transactional","orthogencare.com",              "500–1,000",    "Medium",   "orthogencare.com has dedicated page — build equivalent"),
    ("shoulder pain treatment Kerala",           "Not in Top 10","—",                                         "Transactional","Ayurveda, orthogencare.com",    "1,000–2,000",  "Medium",   "orthogencare.com ranks — Regencare needs a shoulder page"),
    ("anti aging treatment Kerala",              "Not in Top 10","—",                                         "Transactional","Aesthetic clinics",             "2,000–5,000",  "Medium",   "Dermatology opportunity — anti-aging content hub needed"),
    ("PRP for hair growth Kerala",               "Not in Top 10","—",                                         "Transactional","DermaVue, Hair clinics",        "2,000–5,000",  "High",     "High volume — Regencare invisible despite offering service"),
    ("IV therapy Kerala wellness",               "Not in Top 10","—",                                         "Transactional","Ageon, VCare",                  "500–1,000",    "Low",      "Service exists — needs dedicated landing page"),
    ("alopecia treatment Kerala dermatologist",  "Not in Top 10","—",                                         "Transactional","DermaVue, Lybrate, Practo",     "1,000–2,000",  "High",     "Clinical term — Dr. Aswathi profile should target this"),
]

UAE = [
    ("Kerala doctor treatment UAE",                  "Not in Top 10","—","Navigational", "Ayurveda clinics Dubai",         "500–1,000","Medium","Malayalee NRI segment — returning to Kerala for treatment"),
    ("PRP therapy Kerala doctor Dubai",              "Not in Top 10","—","Transactional","Dubai hair clinics",             "500–1,000","Medium","NRI patients search before travelling to Kerala"),
    ("regenerative medicine Kerala visit",           "Not in Top 10","—","Transactional","Not many competitors",           "200–500",  "Medium","Low competition — NRI landing page opportunity"),
    ("stem cell therapy India Kerala",               "Not in Top 10","—","Transactional","Chaitanya, Kokilaben",           "500–1,000","Medium","Medical tourism angle — Regencare can target NRI segment"),
    ("knee pain treatment India Kerala visit",       "Not in Top 10","—","Transactional","Medical tourism portals",        "200–500",  "Medium","NRIs returning for treatment — dedicated NRI page opportunity"),
    ("hair treatment India Kerala visit",            "Not in Top 10","—","Transactional","Cutis International (UAE+Kerala)","200–500", "Medium","Cutis has UAE branch — Regencare needs NRI targeting strategy"),
    ("non surgical treatment Kerala medical tourism","Not in Top 10","—","Transactional","Medical tourism portals",        "200–500",  "Low",  "Long-term — build Kerala medical tourism content"),
    ("orthopedic treatment Kerala India UAE patient","Not in Top 10","—","Transactional","Not many competitors",           "200–500",  "Low",  "Low competition — UAE Malayalee audience"),
    ("regencare appointment UAE patient",            "Not in Top 10","—","Navigational", "—",                              "< 200",    "Low",  "Branded search from UAE — needs WhatsApp booking and NRI page"),
    ("best hospital Kerala joint pain UAE NRI",      "Not in Top 10","—","Transactional","Medical tourism portals",        "200–500",  "Low",  "Medical tourism — low competition, niche NRI audience"),
    ("PRP treatment cost Kerala Dubai comparison",   "Not in Top 10","—","Informational","Dubai clinic sites",             "200–500",  "Medium","NRI comparing cost before travelling — needs cost comparison content"),
    ("Malayalam speaking doctor knee pain UAE",      "Not in Top 10","—","Navigational", "HeliumDoc, Aster UAE",           "200–500",  "Low",  "Malayalam NRI — Dr. Vineeth profile helps"),
    ("regencare Kochi online consultation",          "Not in Top 10","—","Transactional","Practo, Lybrate",                "200–500",  "Medium","NRI or outstation patient wanting online consult before visit"),
]

CONVERSATIONAL = [
    # Questions patients actually type
    ("is PRP treatment safe India",                   "Not in Top 10","—","Question",    "DHI India, Oliva, Apollo",       "5,000–10,000","High",    "Very high volume FAQ — write dedicated PRP safety blog post"),
    ("PRP vs surgery knee pain India",                "Not in Top 10","—","Question",    "Manipal, Alleviate Pain",        "2,000–5,000",  "High",   "High intent — write PRP vs surgery article targeting Kerala"),
    ("which is better PRP or GFC for hair India",     "Not in Top 10","—","Question",    "Multiple comparison sites",      "5,000–10,000","Critical","Extremely high volume — write PRP vs GFC comparison page"),
    ("how to stop hair fall naturally Kochi",         "Not in Top 10","—","Question",    "Oliva, Sanjeevanam",             "2,000–5,000",  "High",   "Blog post drives top-of-funnel patients to GFC/PRP page"),
    ("why is my hair falling out Kochi Kerala",       "Not in Top 10","—","Question",    "Oliva, DermaVue, Zaaya",         "2,000–5,000",  "High",   "Very common patient query — hair loss cause + treatment blog needed"),
    ("knee pain home remedy Kerala",                  "Not in Top 10","—","Question",    "Ayurveda clinics dominate",      "5,000–10,000","High",   "Massive volume — blog post comparing home remedies vs PRP"),
    ("hair fall after pregnancy treatment Kochi",     "Not in Top 10","—","Question",    "Oliva, DermaVue, Twacha",        "1,000–2,000",  "High",   "Specific patient segment — postpartum hair loss blog needed"),
    ("what is regenerative medicine India",           "Not in Top 10","—","Question",    "Wikipedia, Mayo Clinic",         "5,000–10,000","High",   "Top-of-funnel educational query — cornerstone blog opportunity"),
    ("what is PRP therapy how does it work India",    "Not in Top 10","—","Question",    "Apollo, DHI, Oliva",             "5,000–10,000","High",   "Very high volume — PRP explainer article needed on Regencare blog"),
    ("what is stem cell therapy India",               "Not in Top 10","—","Question",    "Apollo, Health portals",         "5,000–10,000","High",   "Informational — write authoritative stem cell explainer article"),
    ("what is GFC treatment for hair India",          "Not in Top 10","—","Question",    "Various comparison sites",       "2,000–5,000",  "High",   "High volume — GFC explainer article targets this perfectly"),
    ("how many PRP sessions needed knee India",       "Not in Top 10","—","Question",    "Various ortho sites",            "1,000–2,000",  "Medium", "FAQ content on treatment page answers this — wins snippet"),
    ("does PRP work for hair loss India",             "Not in Top 10","—","Question",    "DHI, Oliva, DermaVue",           "5,000–10,000","Critical","Extremely high volume — PRP effectiveness blog needed urgently"),
    ("how long does PRP treatment last Kerala",       "Not in Top 10","—","Question",    "Various sites",                  "1,000–2,000",  "Medium", "FAQ content — add to PRP treatment page to capture snippet"),
    ("can stem cell therapy cure arthritis India",    "Not in Top 10","—","Question",    "Health portals, Apollo",         "1,000–2,000",  "High",   "High intent — write stem cell for arthritis article with NMC-compliant framing"),
    # Symptom-based natural searches
    ("knee pain walking stairs treatment Kochi",      "Not in Top 10","—","Symptom",     "Ayurveda, Practo, Lybrate",      "1,000–2,000",  "High",   "Symptom-specific — blog post targeting this query needed"),
    ("knee swelling treatment Kochi doctor",          "Not in Top 10","—","Symptom",     "Ayurveda clinics, Aster",        "1,000–2,000",  "High",   "Symptom-driven — add knee swelling to condition page content"),
    ("hair thinning on top of head treatment Kochi",  "Not in Top 10","—","Symptom",     "Hair specialty clinics",         "1,000–2,000",  "High",   "Specific symptom — write targeted hair thinning blog post"),
    ("sudden hair fall reason treatment Kerala",      "Not in Top 10","—","Symptom",     "Oliva, Ayurveda sites",          "2,000–5,000",  "High",   "Top-of-funnel symptom search — educational blog needed"),
    ("back of knee pain treatment Kochi",             "Not in Top 10","—","Symptom",     "Ayurveda, Practo",               "500–1,000",    "Medium", "Specific symptom — add to knee condition page FAQ"),
    ("heel pain treatment Kochi Kerala",              "Not in Top 10","—","Symptom",     "orthogencare.com, Ayurveda",     "1,000–2,000",  "High",   "Plantar fasciitis often searched as heel pain — optimise page"),
    ("elbow pain treatment Kochi without surgery",    "Not in Top 10","—","Symptom",     "orthogencare.com",               "500–1,000",    "Medium", "Tennis elbow — orthogencare.com has page, Regencare does not"),
    ("neck pain treatment Kochi without surgery",     "Not in Top 10","—","Symptom",     "Maana Health, Kumar Centre",     "1,000–2,000",  "Medium", "High volume — create neck and spine condition page"),
    ("ankle pain treatment Kochi Kerala",             "Not in Top 10","—","Symptom",     "Practo, Physiotherapy clinics",  "500–1,000",    "Medium", "Sports and injury related — add to sports medicine page"),
    ("skin pigmentation treatment Kochi Kerala",      "Not in Top 10","—","Symptom",     "Skinessence, DermaVue, Aster",   "2,000–5,000",  "Medium", "Dermatology gap — add pigmentation content to skin pages"),
    # Cost-related
    ("PRP treatment cost Kerala price",               "Listed",       "regencare.in/prp-procedure","Transactional","Oliva, Hair O Craft, DermaVue","5,000–10,000","Critical","Cost page exists but not ranking top 3 — add clear cost section"),
    ("GFC therapy cost Kerala price",                 "Not in Top 10","—","Transactional","Hair Wellness, DermaVue",        "2,000–5,000",  "High",   "High volume — add clear pricing section to GFC page"),
    ("stem cell therapy cost India Kerala",           "Not in Top 10","—","Transactional","National portals dominate",      "2,000–5,000",  "High",   "Patients compare costs before booking — add cost FAQ to stem cell page"),
    ("hair loss treatment cost Kochi",                "Not in Top 10","—","Transactional","Oliva, DermaVue, Pristyn",       "2,000–5,000",  "High",   "Very high intent — patients ready to book, need pricing info"),
    ("PRP injection knee cost Kerala",                "Not in Top 10","—","Transactional","orthogencare.com, Practo",       "1,000–2,000",  "High",   "High intent — add PRP knee cost section to treatment page"),
    ("acne treatment cost Kochi Kerala",              "Not in Top 10","—","Transactional","Oliva, Pristyn, Skinessence",    "2,000–5,000",  "Medium", "Dermatology cost query — add pricing to acne treatment page"),
    # Comparison keywords
    ("PRP vs stem cell therapy India",                "Not in Top 10","—","Comparison",  "Health portals, Apollo",         "2,000–5,000",  "High",   "Patients comparing options — write comparison article on blog"),
    ("stem cell vs PRP knee India",                   "Not in Top 10","—","Comparison",  "Alleviate Pain, Orthobiologix",  "1,000–2,000",  "High",   "High intent comparison — blog article drives traffic to both pages"),
    ("knee replacement vs PRP Kerala",                "Not in Top 10","—","Comparison",  "Hospital brands, Manipal",       "1,000–2,000",  "High",   "Core Regencare audience — patients avoiding surgery"),
    ("hair transplant vs PRP Kochi Kerala",           "Not in Top 10","—","Comparison",  "Hair transplant clinics",        "2,000–5,000",  "High",   "Comparison content — Regencare can win on non-surgical angle"),
    ("hair transplant vs GFC which is better Kerala", "Not in Top 10","—","Comparison",  "DHI, La Densitae",               "1,000–2,000",  "High",   "Patients weighing options — comparison blog drives bookings"),
    # Long-tail informational
    ("regenerative medicine for athletes Kerala",     "Not in Top 10","—","Informational","RegenOrthoSport",               "200–500",      "Medium", "Sports niche — Regencare treats athletes, no sports page exists"),
    ("non surgical joint pain clinic near me Kerala", "Not in Top 10","—","Transactional","Maana Health, Epione",          "1,000–2,000",  "High",   "'Near me' high intent — optimise GBP and add LocalBusiness schema"),
    ("stem cell therapy for knee cartilage India",    "Not in Top 10","—","Informational","Apollo, Shri Bone & Joint",      "1,000–2,000",  "High",   "Specific clinical query — write cartilage regeneration article"),
    ("PRP therapy for sports injury India",           "Not in Top 10","—","Informational","Manipal, Max Hospital",          "1,000–2,000",  "High",   "High authority sites rank — Regencare can rank with Kerala angle"),
    ("male pattern baldness treatment Kerala",        "Not in Top 10","—","Transactional","DHI, Hair transplant clinics",  "1,000–2,000",  "High",   "Regencare offers PRP/GFC for this — no dedicated page exists"),
    ("female hair loss treatment Kochi Kerala",       "Not in Top 10","—","Transactional","DermaVue, Oliva, Zaaya",        "2,000–5,000",  "High",   "Women hair loss growing segment — dedicated women's hair page needed"),
    ("hair fall stress treatment Kochi",              "Not in Top 10","—","Informational","Oliva, Twacha, DermaVue",        "1,000–2,000",  "Medium", "Stress hair loss is common — blog post targets this cause"),
    ("hair baldness treatment men Kochi",             "Not in Top 10","—","Transactional","DermaVue, Oliva, Hair clinics",  "1,000–2,000",  "High",   "Male pattern baldness — dedicated men's hair loss page needed"),
]

# ═══════════════════════════════════════════════════════════════
# COVER SHEET
# ═══════════════════════════════════════════════════════════════
ws_cover = wb.active
ws_cover.title = "COVER"
ws_cover.sheet_properties.tabColor = NAVY
ws_cover.sheet_view.showGridLines = False

def cv(ws, r, c, v, fc=WHITE, sz=11, bold=False, italic=False, h="center", bg=None, merge=None, row_h=None):
    if merge:
        ws.merge_cells(merge)
    cl = ws.cell(r, c, v)
    cl.font = Font(name="Calibri", color=fc, size=sz, bold=bold, italic=italic)
    cl.alignment = Alignment(horizontal=h, vertical="center", wrap_text=True)
    if bg: cl.fill = PatternFill("solid", fgColor=bg)
    if row_h: ws.row_dimensions[r].height = row_h
    return cl

# Background — full navy
for row in range(1, 45):
    for col in range(1, 12):
        ws_cover.cell(row, col).fill = PatternFill("solid", fgColor=NAVY)

for i in range(1, 12):
    ws_cover.column_dimensions[get_column_letter(i)].width = 12

ws_cover.row_dimensions[1].height  = 20
ws_cover.row_dimensions[2].height  = 12
ws_cover.row_dimensions[3].height  = 60
ws_cover.row_dimensions[4].height  = 40
ws_cover.row_dimensions[5].height  = 18
ws_cover.row_dimensions[6].height  = 30
ws_cover.row_dimensions[7].height  = 22
ws_cover.row_dimensions[8].height  = 22
ws_cover.row_dimensions[9].height  = 22
ws_cover.row_dimensions[10].height = 22
ws_cover.row_dimensions[11].height = 22
ws_cover.row_dimensions[12].height = 22
ws_cover.row_dimensions[13].height = 18
ws_cover.row_dimensions[14].height = 30
ws_cover.row_dimensions[15].height = 22
ws_cover.row_dimensions[16].height = 22
ws_cover.row_dimensions[17].height = 22
ws_cover.row_dimensions[18].height = 22
ws_cover.row_dimensions[19].height = 22
ws_cover.row_dimensions[20].height = 22
ws_cover.row_dimensions[21].height = 18
ws_cover.row_dimensions[22].height = 30
ws_cover.row_dimensions[23].height = 30
ws_cover.row_dimensions[24].height = 30
ws_cover.row_dimensions[25].height = 30
ws_cover.row_dimensions[26].height = 30
ws_cover.row_dimensions[27].height = 30
ws_cover.row_dimensions[28].height = 18
ws_cover.row_dimensions[29].height = 30
ws_cover.row_dimensions[30].height = 25
ws_cover.row_dimensions[31].height = 20

# Gold top bar
ws_cover.merge_cells("A1:K2")
ws_cover["A1"].fill = PatternFill("solid", fgColor=GOLD)

# Main title
ws_cover.merge_cells("A3:K3")
cv(ws_cover, 3, 1, "REGENCARE.IN", WHITE, 36, bold=True, h="center")

ws_cover.merge_cells("A4:K4")
cv(ws_cover, 4, 1, "Digital Baseline Keyword & SEO Report", "D6EAF8", 18, h="center")

ws_cover.merge_cells("A5:K5")
cv(ws_cover, 5, 1, "Before Snapshot  ·  April 2026  ·  Frozen for Month 6 Comparison", GOLD, 11, italic=True, h="center")

# Divider
ws_cover.merge_cells("A6:K6")
ws_cover["A6"].fill = PatternFill("solid", fgColor=TEAL)

# Section: Prepared by
ws_cover.merge_cells("A7:K7")
cv(ws_cover, 7, 1, "PREPARED BY", GOLD, 9, bold=True, h="center")

ws_cover.merge_cells("A8:K8")
cv(ws_cover, 8, 1, "[Your Company Name]", WHITE, 13, bold=True, h="center")

ws_cover.merge_cells("A9:K9")
cv(ws_cover, 9, 1, "Date: April 2026", "D6EAF8", 10, h="center")

ws_cover.merge_cells("A10:K10")
cv(ws_cover, 10, 1, "Prepared for: Regencare — South India's First Dedicated Regenerative Medicine Centre", "D6EAF8", 10, h="center")

ws_cover.merge_cells("A11:K11")
cv(ws_cover, 11, 1, "Branches: Kochi  |  Calicut  |  Chennai", "D6EAF8", 10, h="center")

ws_cover.merge_cells("A12:K12")
cv(ws_cover, 12, 1, "Website: regencare.in", "D6EAF8", 10, h="center")

# Divider
ws_cover.merge_cells("A13:K13")
ws_cover["A13"].fill = PatternFill("solid", fgColor=TEAL)

# Section: What is inside
ws_cover.merge_cells("A14:K14")
cv(ws_cover, 14, 1, "WHAT IS INSIDE THIS REPORT", GOLD, 9, bold=True, h="center")

sheets_info = [
    ("Kochi",                   "31 keywords",  "All Kochi location — service, condition, and doctor keywords"),
    ("Calicut",                 "16 keywords",  "All Calicut / Kozhikode location keywords"),
    ("Chennai",                 "16 keywords",  "All Chennai location and service keywords"),
    ("Kerala Regional",         "21 keywords",  "State-wide Kerala search terms"),
    ("UAE",                     "13 keywords",  "NRI / Dubai / Medical tourism keywords"),
    ("Conversational & Cost",   "45 keywords",  "Question-based, symptom, cost & comparison keywords"),
]
for i, (sh, cnt, desc) in enumerate(sheets_info, 15):
    ws_cover.merge_cells(f"A{i}:C{i}")
    cv(ws_cover, i, 1, sh, GOLD, 10, bold=True, h="left")
    ws_cover.merge_cells(f"D{i}:E{i}")
    cv(ws_cover, i, 4, cnt, "D6EAF8", 10, h="center")
    ws_cover.merge_cells(f"F{i}:K{i}")
    cv(ws_cover, i, 6, desc, WHITE, 10, h="left")

# Divider
ws_cover.merge_cells("A21:K21")
ws_cover["A21"].fill = PatternFill("solid", fgColor=TEAL)

# Scorecard
ws_cover.merge_cells("A22:K22")
cv(ws_cover, 22, 1, "CURRENT SITE SCORECARD — BEFORE BUILD", GOLD, 9, bold=True, h="center")

scores = [
    ("Technical SEO",      "2 / 10", "7 H1 tags, empty sitemap, blog 404, zero schema across all pages"),
    ("Keyword Rankings",   "5 / 10", "Strong for brand terms — absent from PRP hair, GFC, sports, back pain"),
    ("Brand Protection",   "2 / 10", "orthogencare.com outranking Regencare for Dr. Vineeth MB"),
    ("GEO / AI Engines",   "1 / 10", "Not cited in ChatGPT, Gemini, or Perplexity — zero schema"),
    ("Local SEO",          "4 / 10", "Good reviews (4.8★) — GBP incomplete, Chennai email mismatch"),
    ("Content",            "2 / 10", "1 blog post, no research articles, no complete doctor profiles"),
    ("Patient Conversion", "1 / 10", "No booking system, no chatbot, no callback — contact form only"),
]
for i, (area, score, note) in enumerate(scores, 23):
    ws_cover.merge_cells(f"A{i}:C{i}")
    cv(ws_cover, i, 1, area, WHITE, 10, bold=True, h="left")
    ws_cover.merge_cells(f"D{i}:E{i}")
    s_cell = ws_cover.cell(i, 4, score)
    s_cell.font = Font(name="Calibri", color=RED, size=11, bold=True)
    s_cell.alignment = Alignment(horizontal="center", vertical="center")
    s_cell.fill = PatternFill("solid", fgColor=RED_LT)
    ws_cover.merge_cells(f"F{i}:K{i}")
    cv(ws_cover, i, 6, note, "D6EAF8", 9, h="left")

# Divider
ws_cover.merge_cells("A30:K30")
ws_cover["A30"].fill = PatternFill("solid", fgColor=TEAL)

# Footer
ws_cover.merge_cells("A31:K31")
cv(ws_cover, 31, 1,
   "CONFIDENTIAL  ·  Prepared for Regencare  ·  Not for public distribution  ·  April 2026",
   GOLD, 8, italic=True, h="center")

# Gold bottom bar
ws_cover.merge_cells("A32:K33")
ws_cover["A32"].fill = PatternFill("solid", fgColor=GOLD)

# ── Build all data sheets ────────────────────────────────────────
build_sheet(wb, "Kochi",                 "1F6F8B", KOCHI)
build_sheet(wb, "Calicut",               "1E8449", CALICUT)
build_sheet(wb, "Chennai",               "C0392B", CHENNAI)
build_sheet(wb, "Kerala Regional",       "D4850A", KERALA_REGIONAL)
build_sheet(wb, "UAE",                   "8E44AD", UAE)
build_sheet(wb, "Conversational & Cost", "117A65", CONVERSATIONAL)

# ── MASTER SUMMARY sheet ────────────────────────────────────────
ws_sum = wb.create_sheet("MASTER SUMMARY")
ws_sum.sheet_properties.tabColor = NAVY
ws_sum.sheet_view.showGridLines = False

for row in range(1, 3):
    ws_sum.merge_cells(f"A{row}:G{row}")
ws_sum.merge_cells("A1:G1")
ws_sum["A1"].value = "REGENCARE.IN  ·  MASTER KEYWORD SUMMARY  ·  April 2026"
ws_sum["A1"].fill = PatternFill("solid", fgColor=NAVY)
ws_sum["A1"].font = Font(name="Calibri", color=WHITE, size=14, bold=True)
ws_sum["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_sum.row_dimensions[1].height = 36

ws_sum.merge_cells("A2:G2")
ws_sum["A2"].value = "Prepared for: Regencare  ·  Confidential  ·  Before-build snapshot"
ws_sum["A2"].fill = PatternFill("solid", fgColor=TEAL_LIGHT)
ws_sum["A2"].font = Font(name="Calibri", color=NAVY, size=10, italic=True)
ws_sum["A2"].alignment = Alignment(horizontal="center", vertical="center")
ws_sum.row_dimensions[2].height = 20

# Table header
hdr_vals = ["Location / Category", "Total Keywords", "Ranking #1–#3", "Ranking #4–#10", "Not in Top 10", "Critical Gaps", "Status"]
hdr_row(ws_sum, 4, hdr_vals)
ws_sum.row_dimensions[4].height = 26

all_data = [
    ("Kochi",                  KOCHI),
    ("Calicut",                CALICUT),
    ("Chennai",                CHENNAI),
    ("Kerala Regional",        KERALA_REGIONAL),
    ("UAE",                    UAE),
    ("Conversational & Cost",  CONVERSATIONAL),
]

for i, (loc, data) in enumerate(all_data, 5):
    rw = sum(1 for r in data if str(r[1]) in ("#1","#2","#3"))
    im = sum(1 for r in data if any(str(r[1])==f"#{j}" for j in range(4,11)))
    nt = sum(1 for r in data if "Not" in str(r[1]) or "Listed" in str(r[1]))
    cr = sum(1 for r in data if str(r[6]) == "Critical")
    status = "Needs Immediate Action" if cr >= 3 else "Action Required" if nt > 5 else "Monitor"
    row_vals = [loc, len(data), rw, im, nt, cr, status]
    for c, val in enumerate(row_vals, 1):
        cl = ws_sum.cell(i, c, val)
        cl.border = bdr()
        cl.alignment = Alignment(horizontal="center", vertical="center")
        bg = GREY if i % 2 == 0 else WHITE
        cl.fill = PatternFill("solid", fgColor=bg)
        cl.font = Font(name="Calibri", size=10)
        if c == 1:
            cl.font = Font(name="Calibri", size=10, bold=True, color=NAVY)
            cl.alignment = Alignment(horizontal="left", vertical="center")
        if c == 3:
            cl.fill = PatternFill("solid", fgColor=GREEN_LT)
            cl.font = Font(name="Calibri", bold=True, color=GREEN, size=10)
        if c == 5:
            cl.fill = PatternFill("solid", fgColor=RED_LT)
            cl.font = Font(name="Calibri", bold=True, color=RED, size=10)
        if c == 6 and val > 0:
            cl.fill = PatternFill("solid", fgColor=RED_LT)
            cl.font = Font(name="Calibri", bold=True, color=RED, size=10)
        if c == 7:
            fc = RED if "Immediate" in str(val) else AMBER if "Required" in str(val) else GREEN
            bg2 = RED_LT if "Immediate" in str(val) else AMBER_LT if "Required" in str(val) else GREEN_LT
            cl.fill = PatternFill("solid", fgColor=bg2)
            cl.font = Font(name="Calibri", bold=True, color=fc, size=10)
    ws_sum.row_dimensions[i].height = 22

# Totals
tr = 5 + len(all_data)
all_flat = [r for _, d in all_data for r in d]
t_rw = sum(1 for r in all_flat if str(r[1]) in ("#1","#2","#3"))
t_im = sum(1 for r in all_flat if any(str(r[1])==f"#{j}" for j in range(4,11)))
t_nt = sum(1 for r in all_flat if "Not" in str(r[1]) or "Listed" in str(r[1]))
t_cr = sum(1 for r in all_flat if str(r[6]) == "Critical")
for c, val in enumerate(["TOTAL", len(all_flat), t_rw, t_im, t_nt, t_cr, ""], 1):
    cl = ws_sum.cell(tr, c, val)
    cl.fill = PatternFill("solid", fgColor=NAVY)
    cl.font = Font(name="Calibri", bold=True, color=WHITE, size=11)
    cl.alignment = Alignment(horizontal="center", vertical="center")
    cl.border = bdr()
ws_sum.row_dimensions[tr].height = 26

# Column widths
for i, w in enumerate([28, 14, 14, 14, 14, 14, 22], 1):
    ws_sum.column_dimensions[get_column_letter(i)].width = w

# ── Save ────────────────────────────────────────────────────────
out = r"c:\project\AI RESEARCH INTELLIGENCE SYSTEM\Regencare_Baseline_Keyword_Report.xlsx"
wb.save(out)
print(f"Saved: {out}")
