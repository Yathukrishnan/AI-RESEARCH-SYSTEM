"""
Regencare Competitive Study Excel v2
Adds major hospital sites + SEO strategy decoded + content structure + full blueprint
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re

wb = openpyxl.Workbook()

# ── PALETTE ──────────────────────────────────────────────────────────────────
NAVY   = "0D2B55"; GOLD   = "C9A02D"; WHITE  = "FFFFFF"; LIGHT  = "F4F6F9"
MID    = "D0D8E4"; TEXT   = "1A1A2E"; RED    = "C0392B"; RED_L  = "FDECEA"
GREEN  = "1A7A4A"; GRN_L  = "E8F5EE"; AMB    = "D68A00"; AMB_L  = "FFF3CD"
TEAL   = "0D6E6E"; TEA_L  = "E0F4F4"; GREY   = "6C757D"; GRY_L  = "F0F0F0"
PURP   = "6B21A8"; PUR_L  = "F3E8FF"; BLUE   = "1E3A8A"; BLU_L  = "DBEAFE"
CRIT   = "7F1D1D"

def F(h): return PatternFill("solid", fgColor=h)
def Ft(h=TEXT, bold=False, sz=10, italic=False):
    return Font(color=h, bold=bold, size=sz, italic=italic, name="Calibri")
def Al(h="left", v="center", w=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=w)
def Br(c=MID):
    s = Side(style="thin", color=c)
    return Border(left=s, right=s, top=s, bottom=s)
def cw(ws, col, w): ws.column_dimensions[get_column_letter(col)].width = w
def rh(ws, row, h): ws.row_dimensions[row].height = h

def C(ws, r, c, val, bg=WHITE, fg=TEXT, bold=False, sz=10, ha="left", wrap=True):
    x = ws.cell(r, c, val)
    x.fill=F(bg); x.font=Ft(fg,bold,sz); x.alignment=Al(ha,"center",wrap)
    x.border=Br(); return x

def H(ws, r, c, val, bg=NAVY, fg=WHITE, bold=True, sz=10, ha="center"):
    return C(ws,r,c,val,bg,fg,bold,sz,ha)

def MH(ws,r1,c1,r2,c2,val,bg=NAVY,fg=WHITE,sz=11,bold=True):
    ws.merge_cells(start_row=r1,start_column=c1,end_row=r2,end_column=c2)
    x=ws.cell(r1,c1,val); x.fill=F(bg); x.font=Ft(fg,bold,sz)
    x.alignment=Al("center","center",True); x.border=Br(); return x

def TC(ws,r,c,val):
    """✓/✗/~ colour cell"""
    m={"Yes":(GRN_L,GREEN,"✓ Yes"),"No":(RED_L,RED,"✗ No"),
       "Partial":(AMB_L,AMB,"~ Partial"),"Unknown":(GRY_L,GREY,"? Unknown"),
       "Strong":(GRN_L,GREEN,"✓ Strong"),"Weak":(AMB_L,AMB,"~ Weak"),
       "None":(RED_L,RED,"✗ None"),"High":(GRN_L,GREEN,"High"),
       "Medium":(AMB_L,AMB,"Medium"),"Low":(RED_L,RED,"Low"),
       "Very High":(GRN_L,GREEN,"Very High"),"Very Low":(RED_L,RED,"Very Low")}
    bg,fg,txt=m.get(val,(WHITE,TEXT,val))
    x=ws.cell(r,c,txt); x.fill=F(bg); x.font=Ft(fg,True,10)
    x.alignment=Al("center","center",False); x.border=Br(); return x

def PC(ws,r,c,pos):
    """Position colour cell"""
    if str(pos) in("—","NR","Not Ranking","Not in Top 10"):
        bg,fg,v=RED_L,RED,"Not Ranking"
    elif str(pos)=="Regencare":
        bg,fg,v=TEA_L,TEAL,"Regencare"
    else:
        try:
            n=int(re.sub(r'\D','',str(pos)))
            if n<=3: bg,fg,v=GRN_L,GREEN,f"#{n}"
            elif n<=7: bg,fg,v=AMB_L,AMB,f"#{n}"
            else: bg,fg,v=RED_L,RED,f"#{n}"
        except: bg,fg,v=GRY_L,GREY,str(pos)
    x=ws.cell(r,c,v); x.fill=F(bg); x.font=Ft(fg,True,10)
    x.alignment=Al("center","center",False); x.border=Br(); return x

def prio_cell(ws,r,c,p):
    m={"CRITICAL":(RED_L,RED),"HIGH":(AMB_L,AMB),"MEDIUM":(TEA_L,TEAL),"LOW":(GRY_L,GREY)}
    bg,fg=m.get(p,(WHITE,TEXT))
    x=ws.cell(r,c,p); x.fill=F(bg); x.font=Ft(fg,True,9)
    x.alignment=Al("center","center",False); x.border=Br(); return x

def section_title(ws,r,c1,c2,text,bg=NAVY,fg=GOLD,sz=14):
    ws.merge_cells(start_row=r,start_column=c1,end_row=r,end_column=c2)
    x=ws.cell(r,c1,text); x.fill=F(bg); x.font=Ft(fg,True,sz)
    x.alignment=Al("center","center",False); rh(ws,r,32)
def sub_title(ws,r,c1,c2,text,bg=NAVY,fg=WHITE,sz=10):
    ws.merge_cells(start_row=r,start_column=c1,end_row=r,end_column=c2)
    x=ws.cell(r,c1,text); x.fill=F(bg); x.font=Ft(fg,False,sz,italic=True)
    x.alignment=Al("center","center",False); rh(ws,r,20)
def sec_hdr(ws,r,c1,c2,text,bg=NAVY,fg=WHITE,sz=11):
    ws.merge_cells(start_row=r,start_column=c1,end_row=r,end_column=c2)
    x=ws.cell(r,c1,text); x.fill=F(bg); x.font=Ft(fg,True,sz)
    x.alignment=Al("left","center",False); rh(ws,r,26)

# ═══════════════════════════════════════════════════════════════════════════
# DATA DEFINITIONS
# ═══════════════════════════════════════════════════════════════════════════

LOCAL_COMPS = [
    (1,"Orthogen Care","orthogencare.com","Brand Cannibalization","CRITICAL","Ernakulam/Kochi","Ortho PRP / Regenerative Medicine"),
    (2,"DermaVue","dermavue.com","Kerala Clinic Chain","HIGH","7 clinics Kerala+TN","GFC, PRP, Hair Loss"),
    (3,"Oliva Clinic","olivaclinic.com","National Hair Chain","HIGH","Pan-India","Hair Fall, PRP, Transplant"),
    (4,"Cutis International","cutisinternational.com","International Chain","HIGH","Kerala+UAE+UK","Hair Transplant, Skin, Surgery"),
    (5,"La Densitae","ladensitae.com","Hair Chain","MED","13 branches India+Dubai","Hair Transplant, PRP"),
    (6,"Zaaya Skin Clinic","zaayaskinclinic.com","Local Kochi","MED","Kochi (3 locations)","GFC, PRP, Hair Loss, Skin"),
    (7,"DHI International","dhiinternational.com","Global Hair Brand","MED","Global — Kochi branch","Hair Transplant (DHI)"),
    (8,"DH Clinic","dh-clinic.com","Ortho+Stem Cell","MED","Kerala (multi-branch)","Ortho, Stem Cell, PRP (Surecell AU)"),
    (9,"Hair Wellness Clinic","hairwellnessclinic.com","Hair Specialist","MED","Kerala","GFC, PRP, Hair treatments"),
    (10,"Epione Pain Centre","paincentre.in","Pain & Regen","MED","Kerala","PRP, Stem Cell, Pain"),
]

HOSPITAL_COMPS = [
    # id, name, domain, type, da, locations, stem_cell_focus, word_ct, h2_count, schema, faq, doctor_cred, backlinks, eeat
    (11,"Apollo Hospitals","apollohospitals.com","National Hospital Chain",72,"Pan-India + International","Regrow Cell Therapy — DCGI-approved formulations (OSSGROW, CARTIGROW) for cartilage and bone",1300,12,"Strong","Yes","Named specialists with profiles",  "Very High","Very High"),
    (12,"Apollo Spectra","apollospectra.com","Hospital (Apollo Group)",65,"Pan-India, Tier 2 cities","Regrow Stem Cell Therapy for Bones and Cartilage — proprietary branded treatment",1300,12,"Strong","Yes","Linked specialist profiles","Very High","Very High"),
    (13,"Manipal Hospitals","manipalhospitals.com","National Hospital Chain",70,"Pan-India + International","Blog: Stem Cell Therapy in Orthopaedics — educational, research-backed, clinical tone",1300,7,"Strong","Yes","Dr. Lokesh A Veerappa — Consultant Orthopaedic & Robotic Joint Replacement","Very High","Very High"),
    (14,"Kokilaben Hospital","kokilabenhospital.com","Super-Speciality Hospital",62,"Mumbai (National reach)","Comprehensive PRP page — 2,100 words, 9 H2 sections, treatment comparison table",2100,9,"Strong","Yes","Named rehabilitation physicians, 10+ years experience","High","Very High"),
    (15,"Stem Cell Care India","stemcellcareindia.com","Dedicated Stem Cell Centre",44,"Delhi (International patients)","8,000+ word homepage, 50+ conditions, international patient guide, GMP lab, NABH",8000,9,"Strong","Yes","Stem Cell Scientists and Medical Team (not individually named)","Medium","High"),
    (16,"RegenOrthoSport","regenorthosport.in","Regen Ortho Clinic",32,"Hyderabad, Mumbai, Bangalore, Dallas","Stem Cell + HCP Therapy for knee — 1,800 words, 4 named doctors, IOF member",1800,12,"Partial","Yes","Dr. Venkatesh Movva + 3 specialists","Medium","High"),
    (17,"Chaitanya Stem Cell","chaitanyastemcell.com","Dedicated Stem Cell Centre",28,"Pune (National reach)","3,500-word OA treatment page — academic citations (6 PubMed), condition-specific structure",3500,7,"Partial","Yes","Dr. Anant Bagul — Orthopedician Pune","Low","Medium"),
]

KEYWORDS = [
    ("stem cell therapy Kerala",480,"Commercial","Regen — Kerala"),
    ("stem cell treatment knee India",390,"Commercial","Regen — National"),
    ("regenerative medicine India",720,"Commercial","Regen — National"),
    ("PRP treatment Kerala",1600,"Commercial","PRP — Kerala"),
    ("GFC therapy Kochi",1300,"Commercial","GFC — Kochi"),
    ("hair loss treatment Kochi",2400,"Commercial","Hair — Kochi"),
    ("knee pain treatment without surgery",720,"Commercial","Ortho — Kerala"),
    ("best stem cell hospital India",590,"Transact.","Regen — National"),
    ("stem cell therapy cost India",480,"Transact.","Conversational"),
    ("PRP for knee pain India",320,"Commercial","PRP — National"),
    ("regenerative medicine Kochi",170,"Navigat.","Regen — Kochi"),
    ("stem cell treatment Kochi",140,"Commercial","Regen — Kochi"),
    ("non-surgical knee treatment Kerala",170,"Commercial","Ortho — Kerala"),
    ("GFC therapy cost Kerala",170,"Transact.","Conversational"),
    ("hair transplant Kochi",1900,"Transact.","Hair — Kochi"),
]

# Positions: local comps (1-10), hospital comps (11-17), Regencare
LOCAL_POS = {
    1: ["3","NR","4","NR","NR","NR","4","NR","NR","4","4","5","4","NR","NR"],
    2: ["NR","NR","NR","4","1","2","NR","NR","NR","NR","NR","NR","NR","NR","NR"],
    3: ["NR","NR","NR","2","NR","1","NR","NR","NR","NR","NR","NR","NR","NR","NR"],
    4: ["NR","NR","NR","NR","7","NR","NR","NR","NR","NR","NR","NR","NR","NR","3"],
    5: ["NR","NR","NR","NR","5","4","NR","NR","NR","NR","NR","NR","NR","NR","2"],
    6: ["NR","NR","NR","NR","3","5","NR","NR","NR","NR","NR","NR","NR","NR","NR"],
    7: ["NR","NR","NR","NR","NR","NR","NR","NR","NR","NR","NR","NR","NR","NR","1"],
    8: ["5","NR","NR","NR","NR","NR","5","NR","NR","5","NR","5","6","NR","NR"],
    9: ["NR","NR","NR","NR","8","NR","NR","NR","NR","NR","NR","NR","NR","NR","NR"],
   10: ["7","NR","NR","NR","NR","NR","7","NR","NR","7","NR","7","8","NR","NR"],
}
HOSP_POS = {
   11: ["NR","3","2","NR","NR","NR","NR","2","3","3","NR","NR","NR","NR","NR"],
   12: ["NR","4","3","NR","NR","NR","NR","3","4","4","NR","NR","NR","NR","NR"],
   13: ["NR","NR","4","NR","NR","NR","NR","4","5","5","NR","NR","NR","NR","NR"],
   14: ["NR","NR","NR","3","NR","NR","NR","5","NR","2","NR","NR","NR","NR","NR"],
   15: ["NR","2","NR","NR","NR","NR","NR","1","2","NR","NR","NR","NR","NR","NR"],
   16: ["NR","5","5","NR","NR","NR","5","6","6","6","NR","NR","5","NR","NR"],
   17: ["NR","6","NR","NR","NR","NR","NR","7","NR","7","NR","NR","NR","NR","NR"],
}
REGEN_POS=["1","NR","NR","3","9","NR","2","NR","NR","3","2","2","2","NR","NR"]

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 1: COVER
# ═══════════════════════════════════════════════════════════════════════════
ws = wb.active; ws.title="COVER"
ws.sheet_view.showGridLines=False
for c in range(1,13): cw(ws,c,14)

for r in range(1,9):
    for c in range(1,13): ws.cell(r,c).fill=F(NAVY)
    rh(ws,r,22)

ws.merge_cells("A1:L4")
x=ws.cell(1,1,"REGENCARE.IN — COMPETITIVE INTELLIGENCE STUDY  v2.0")
x.fill=F(NAVY);x.font=Ft(GOLD,True,20);x.alignment=Al("center","center",False)

ws.merge_cells("A5:L6")
x=ws.cell(5,1,"LOCAL Clinic Competitors + Major National Hospital Sites  |  SEO Strategy Decoded  |  Page-1 Blueprint")
x.fill=F(NAVY);x.font=Ft(WHITE,False,12,italic=True);x.alignment=Al("center","center",False)

ws.merge_cells("A7:L8")
x=ws.cell(7,1,"Prepared by NT Global Digital  ·  Client: Regencare.in  ·  Version 2.0  ·  May 2026  ·  Confidential")
x.fill=F(NAVY);x.font=Ft(MID,False,10);x.alignment=Al("center","center",False)

rh(ws,9,12)
sec_hdr(ws,10,1,12,"SHEET INDEX — 12 SHEETS IN THIS FILE")
sheets_idx=[
    ("1","COVER","This page — study scope, legend, what's new in v2"),
    ("2","LOCAL COMPETITORS","Original 10 local/clinic competitors — overview, threat tier, notes"),
    ("3","MAJOR HOSPITAL SITES","7 major hospital & national regen sites — deep profile with SEO metrics"),
    ("4","KEYWORD RANKING MAP","15 key keywords — who ranks where: local clinics vs. hospital sites vs. Regencare"),
    ("5","CONTENT STRUCTURE DECODED","How every major competitor structures their pages — H1/H2 map, word count, sections"),
    ("6","SEO STRATEGY MATRIX","Which SEO strategies each competitor uses — schema, E-E-A-T, internal linking, GEO"),
    ("7","ON-PAGE SEO AUDIT","Technical on-page metrics: title, H1, meta, schema, FAQ, booking, NMC compliance"),
    ("8","AUTHORITY & TRUST","Google ratings, review counts, Domain Authority, backlinks, E-E-A-T scores"),
    ("9","SERVICE COVERAGE","Treatments offered: local clinics vs. national hospitals — gap matrix"),
    ("10","WHITE SPACE & GAPS","What nobody owns well — Regencare's fastest page-1 opportunities"),
    ("11","REGENCARE SEO BLUEPRINT","Complete page-1 strategy: site architecture, content template, schema map, CTA flow"),
    ("12","ACTION PRIORITIES","17 ranked actions with competitive evidence, owner, GTM step, expected outcome"),
]
for i,(num,name,desc) in enumerate(sheets_idx):
    r=11+i; rh(ws,r,22); bg=LIGHT if i%2==0 else WHITE
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=1)
    C(ws,r,1,num,bg,NAVY,True,11,"center")
    ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=4)
    C(ws,r,2,name,bg,NAVY,True,10,"left")
    ws.merge_cells(start_row=r,start_column=5,end_row=r,end_column=12)
    C(ws,r,5,desc,bg,TEXT,False,10,"left")

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 2: LOCAL COMPETITORS
# ═══════════════════════════════════════════════════════════════════════════
ws2=wb.create_sheet("LOCAL COMPETITORS")
ws2.sheet_view.showGridLines=False
col_w2=[4,22,30,24,12,32,40]
for i,w in enumerate(col_w2): cw(ws2,i+1,w)

section_title(ws2,1,1,7,"LOCAL & CLINIC COMPETITORS  —  10 Sites Outranking Regencare on Kerala Keywords")
sub_title(ws2,2,1,7,"Sites identified by frequency of appearance on SERPs for 20 highest-volume Regencare keyword clusters  ·  May 2026")

hdrs=["#","Competitor","Domain","Type","Threat","Locations","Focus / Key Threat to Regencare"]
for c,h in enumerate(hdrs,1): H(ws2,3,c,h); rh(ws2,3,32)

NOTES_L={
    1:"SAME DOCTOR (Dr. Vineeth MB) — appears on identical SERPs as regencare.in. #1 priority: no other fix delivers ROI until this is resolved.",
    2:"7 clinics, 4.8★/1,438 reviews. #1 for GFC Kochi. 4,900-word pages with schema. Regencare ranks #9 — content depth is the gap.",
    3:"4.9★/1,475 reviews, 109K+ procedures. Dominates hair loss Kochi SERPs. National chain with 115 MD dermatologists.",
    4:"UAE+UK presence competes directly in Regencare's NRI medical tourism segment. 20+ hair services, celebrity endorsements.",
    5:"13+ branches + Dubai. Ranks for GFC and hair transplant. No named doctors on pages — E-E-A-T weakness Regencare can exploit.",
    6:"Strong local GFC content, #3 for GFC Kochi. No pricing, no named credentials — both are exploitable gaps.",
    7:"Global brand, #1 IMRB for hair transplant. Limited overlap — Regencare USP is ortho regen, not cosmetic transplant.",
    8:"ONLY India clinic with Surecell Australia affiliation for regen medicine — low online authority but unique credentialing claim.",
    9:"Specialist GFC/hair. Low authority, thin content — outranking feasible with one well-structured GFC page.",
    10:"7yr Kerala PRP+Stem Cell+Prolotherapy — direct ortho regen overlap. Low DA. Easy to outrank with targeted content.",
}
threat_c={"CRITICAL":(RED_L,RED),"HIGH":(AMB_L,AMB),"MED":(TEA_L,TEAL)}
for i,(cid,name,domain,ctype,tier,locs,focus) in enumerate(LOCAL_COMPS):
    r=4+i; bg=LIGHT if i%2==0 else WHITE; rh(ws2,r,44)
    C(ws2,r,1,str(cid),bg,NAVY,True,11,"center")
    C(ws2,r,2,name,bg,NAVY,True,10,"left")
    C(ws2,r,3,domain,bg,TEAL,False,10,"left")
    C(ws2,r,4,ctype,bg,TEXT,False,9,"left")
    tb,tf=threat_c.get(tier,(GRY_L,GREY))
    x=ws2.cell(r,5,tier); x.fill=F(tb);x.font=Ft(tf,True,9)
    x.alignment=Al("center","center",False);x.border=Br()
    C(ws2,r,6,locs,bg,TEXT,False,9,"left")
    C(ws2,r,7,NOTES_L.get(cid,focus),bg,GREY,False,9,"left")

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 3: MAJOR HOSPITAL SITES
# ═══════════════════════════════════════════════════════════════════════════
ws3=wb.create_sheet("MAJOR HOSPITAL SITES")
ws3.sheet_view.showGridLines=False
hcols=[4,22,30,6,38,9,8,8,10,10,38,10,10]
for i,w in enumerate(hcols): cw(ws3,i+1,w)

section_title(ws3,1,1,13,"MAJOR NATIONAL HOSPITAL & SPECIALISED REGEN SITES  —  7 Sites Competing on Stem Cell / Regen Keywords")
sub_title(ws3,2,1,13,"These sites rank on page 1 nationally and regionally for stem cell, PRP, and regenerative medicine keywords — often ahead of all local Kerala clinics")

hh=["#","Name","Domain","DA\nEst.","Stem Cell / Regen Content Focus","Word\nCount","H2\nCount","Schema","FAQ","Doctor\nCredentials","Unique SEO Advantage","Backlink\nProfile","E-E-A-T"]
for c,h in enumerate(hh,1): H(ws3,3,c,h,sz=9); rh(ws3,3,40)

HOSP_NOTES={
    11:"DCGI-approved branded treatments (OSSGROW, CARTIGROW) — no other competitor can claim this. Breadcrumb schema + location pages for every city. Brand authority alone earns top-3 positions.",
    12:"Same Apollo brand authority. Dedicated Spectra brand for Tier-2 cities. Strong breadcrumb + MedicalProcedure schema. Treatment structured as: overview→who qualifies→how→benefits→risks→FAQ.",
    13:"Doctor-authored blog with PubMed citation — gold standard E-E-A-T signal. 40+ hospital locations = massive internal linking. Named doctor on every page. Research disclaimers build trust.",
    14:"Strongest PRP page structure found: 9 H2s, treatment comparison table (PRP vs. others), 'Why choose us' section, ultrasound-guided procedure mentioned — premium differentiators.",
    15:"8,000+ word homepage — most comprehensive stem cell content indexed. International patient guide, multi-country pages, GMP lab credentials. 50+ conditions = enormous long-tail keyword coverage.",
    16:"Multi-city presence + Dallas (USA) = cross-border medical tourism SEO. IOF member badge. Named 4 doctors with specialties. Knee-specific HCP Therapy + Stem Cell — two treatment pages for one condition.",
    17:"6 PubMed academic citations on condition page — highest research authority signal found. Named orthopedician. Condition → Symptoms → Treatment → FAQ structure. Easy to replicate for Regencare.",
}
for i,(cid,name,domain,ctype,da,locs,focus,wc,h2,schema,faq,doc,bl,eeat) in enumerate(HOSPITAL_COMPS):
    r=4+i; bg=LIGHT if i%2==0 else WHITE; rh(ws3,r,55)
    C(ws3,r,1,str(cid),bg,NAVY,True,11,"center")
    C(ws3,r,2,name,bg,NAVY,True,10,"left")
    C(ws3,r,3,domain,bg,TEAL,False,9,"left")
    da_bg=GRN_L if da>=60 else AMB_L if da>=35 else RED_L
    da_fg=GREEN if da>=60 else AMB if da>=35 else RED
    x=ws3.cell(r,4,str(da));x.fill=F(da_bg);x.font=Ft(da_fg,True,10)
    x.alignment=Al("center","center",False);x.border=Br()
    C(ws3,r,5,focus,bg,TEXT,False,9,"left")
    wc_bg=GRN_L if wc>=3000 else AMB_L if wc>=1500 else RED_L
    wc_fg=GREEN if wc>=3000 else AMB if wc>=1500 else RED
    x=ws3.cell(r,6,f"{wc:,}");x.fill=F(wc_bg);x.font=Ft(wc_fg,True,10)
    x.alignment=Al("center","center",False);x.border=Br()
    h2_bg=GRN_L if h2>=9 else AMB_L if h2>=6 else RED_L
    h2_fg=GREEN if h2>=9 else AMB if h2>=6 else RED
    x=ws3.cell(r,7,str(h2));x.fill=F(h2_bg);x.font=Ft(h2_fg,True,10)
    x.alignment=Al("center","center",False);x.border=Br()
    TC(ws3,r,8,schema); TC(ws3,r,9,faq)
    C(ws3,r,10,doc,bg,TEXT,False,9,"left")
    C(ws3,r,11,HOSP_NOTES.get(cid,""),bg,GREY,False,9,"left")
    TC(ws3,r,12,bl); TC(ws3,r,13,eeat)

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 4: KEYWORD RANKING MAP (extended)
# ═══════════════════════════════════════════════════════════════════════════
ws4=wb.create_sheet("KEYWORD RANKING MAP")
ws4.sheet_view.showGridLines=False
cw(ws4,1,36); cw(ws4,2,9); cw(ws4,3,12); cw(ws4,4,16); cw(ws4,5,11)
for c in range(6,24): cw(ws4,c,11)

section_title(ws4,1,1,23,"KEYWORD RANKING MAP  —  Regencare vs. 10 Local Clinics + 7 Major Hospital Sites")
sub_title(ws4,2,1,23,"Green=#1-3 · Amber=#4-7 · Red=#8-10 · Not Ranking=no page-1 presence · Positions estimated from live SERP analysis")

local_names=["Orthogen","DermaVue","Oliva","Cutis","La Dens.","Zaaya","DHI","DH Clinic","HairWell","Epione"]
hosp_names=["Apollo Hosp","Apollo Spctr","Manipal","Kokilaben","SCCareIndia","RegenOrtho","Chaitanya"]

# Headers row 3 — no merging, all individual cells
cw(ws4,1,36); cw(ws4,2,9); cw(ws4,3,12); cw(ws4,4,16)
for c in range(5,24): cw(ws4,c,11)
H(ws4,3,1,"Keyword",NAVY,WHITE,True,9,"left")
H(ws4,3,2,"Vol/mo",NAVY,WHITE,True,9)
H(ws4,3,3,"Intent",NAVY,WHITE,True,9)
H(ws4,3,4,"Cluster",NAVY,WHITE,True,9)
H(ws4,3,5,"REGENCARE",TEAL,WHITE,True,9)
for ci,n in enumerate(local_names,6): H(ws4,3,ci,n,NAVY,WHITE,True,9)
for ci,n in enumerate(hosp_names,16): H(ws4,3,ci,n,PURP,WHITE,True,9)
rh(ws4,3,40)

for ki,(kw,vol,intent,cluster) in enumerate(KEYWORDS):
    r=4+ki; bg=LIGHT if ki%2==0 else WHITE; rh(ws4,r,22)
    C(ws4,r,1,kw,bg,TEXT,False,10,"left")
    C(ws4,r,2,f"{vol:,}",bg,NAVY,True,10,"center")
    ic={"Commercial":TEAL,"Transact.":GREEN,"Navigat.":AMB}.get(intent,TEXT)
    x=ws4.cell(r,3,intent);x.fill=F(bg);x.font=Ft(ic,False,9)
    x.alignment=Al("center","center",False);x.border=Br()
    C(ws4,r,4,cluster,bg,TEXT,False,9,"center")
    PC(ws4,r,5,REGEN_POS[ki])
    for ci,cid in enumerate(range(1,11),6): PC(ws4,r,ci,LOCAL_POS[cid][ki])
    for ci,cid in enumerate(range(11,18),16): PC(ws4,r,ci,HOSP_POS[cid][ki])

ws4.freeze_panes="E4"

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 5: CONTENT STRUCTURE DECODED
# ═══════════════════════════════════════════════════════════════════════════
ws5=wb.create_sheet("CONTENT STRUCTURE DECODED")
ws5.sheet_view.showGridLines=False
for i,w in enumerate([26,18,14,10,10,9,10,14,40],1): cw(ws5,i,w)

section_title(ws5,1,1,9,"CONTENT STRUCTURE DECODED  —  How Top Competitors Build Pages That Rank")
sub_title(ws5,2,1,9,"Analysed via direct page fetch · H2 = heading sections · Word count = total page content · Schema = structured data types implemented")
rh(ws5,2,20)

H(ws5,3,1,"Competitor / Site",NAVY,WHITE,True,10,"left")
H(ws5,3,2,"H1 Pattern",NAVY,WHITE,True,9)
H(ws5,3,3,"Word Count",NAVY,WHITE,True,9)
H(ws5,3,4,"H2 Count",NAVY,WHITE,True,9)
H(ws5,3,5,"FAQ\nSection",NAVY,WHITE,True,9)
H(ws5,3,6,"Pricing\nListed",NAVY,WHITE,True,9)
H(ws5,3,7,"Doctor\nAuthored",NAVY,WHITE,True,9)
H(ws5,3,8,"Academic\nCitations",NAVY,WHITE,True,9)
H(ws5,3,9,"H2 Heading Structure (actual headings found)",NAVY,WHITE,True,9,"left")
rh(ws5,3,40)

CONTENT_DATA=[
    # (name, h1_pattern, wc, h2_ct, faq, pricing, doc_authored, citations, h2_list)
    ("REGENCARE.IN (Current)",
     "Brand/Treatment name — inconsistent",
     3200,4,"No","Partial","Yes","No",
     "GFC Therapy | PRP | About Us | Contact (no condition/treatment depth)"),
    ("Apollo Spectra — Regrow",
     "[Treatment Name] + [Hospital Brand]",
     1300,12,"Yes","No","Linked","No",
     "Overview · What is Regrow? · Who qualifies? · Why conducted? · How implemented? · Benefits · Risks · Symptoms · Treatments · Specialities · Cities · Book Appointment"),
    ("Manipal Hospitals — Stem Cell Blog",
     "Question format: 'Is [Topic] A Breakthrough?'",
     1300,7,"Yes","No","Yes — named MD","PubMed",
     "What are Stem Cells? · Characteristics · What is Stem Cell Therapy? · Therapeutic Use in Orthopaedics · Conditions Treated · Role of Mesenchymal Cells · FAQs"),
    ("Kokilaben Hospital — PRP",
     "[Treatment] (plain, clean)",
     2100,9,"Yes","No","Named physician","No",
     "What is PRP? · How does it work? · Conditions Treated · Benefits · PRP vs. Other Treatments · What to expect · Is it right for you? · Why Choose Kokilaben? · FAQs"),
    ("Stem Cell Care India",
     "[Treatment] in [Country]",
     8000,9,"Yes","No","Team reference","No",
     "Why Choose Us · Treat Your Diseases · Conditions Treated · Exosome Therapies · International Patient Guide · FAQs · Key Benefits"),
    ("RegenOrthoSport — Knee Stem Cell",
     "Benefit-led: 'Breakthrough Non-Surgical Knee Treatments'",
     1800,12,"Yes","No","4 named MDs","No",
     "Regen Therapies for Pain · Commonly Treated Conditions · About Expertise · Non-Surgical Approach · Osteoarthritis · Meniscus Tears · ACL Tears · How Procedures Work · HCP Therapy · Stem Cell Therapy · FAQs"),
    ("Chaitanya Stem Cell — Osteoarthritis",
     "[Condition] Treatment in [Country]",
     3500,7,"Yes","No","Named orthopedician","6 PubMed",
     "Stem Cell for Knee/Hip/Joints · What is OA? · OA Types · Signs and Symptoms · Affected Joints · Patient Testimonials · FAQs"),
    ("DermaVue — GFC/PRP Kochi",
     "[Treatment] Near [City] — Benefit Claim",
     4900,8,"No","Yes","Named MD DVL","Clinical meta-analysis",
     "Why GFC? · What is GFC vs PRP? · Our Protocol · Before & After · Session Guide · Cost Guide · Book Consultation · Why DermaVue?"),
    ("Oliva Clinic — Hair Fall Kochi",
     "[Condition] In [City]: Cost, Procedure, Results & Reviews",
     3200,7,"Yes","Yes","4 named MDs","No",
     "Types of Hair Loss · V-Discover Process · Treatment Options · Cost · Results Timeline · Reviews · FAQs"),
    ("Kokilaben — PRP vs Others",
     "(Comparison table embedded in PRP page)",
     2100,9,"Yes","No","Named","No",
     "PRP vs Steroid Injection vs Surgery vs Hyaluronic Acid — table format showing superiority without making prohibited claims"),
]
for i,(name,h1,wc,h2,faq,pricing,doc,cit,h2_list) in enumerate(CONTENT_DATA):
    r=4+i; bg=TEA_L if i==0 else (LIGHT if i%2==1 else WHITE); rh(ws5,r,50)
    C(ws5,r,1,name,bg,NAVY if i>0 else TEAL,True,10,"left")
    C(ws5,r,2,h1,bg,TEXT,False,9,"left")
    wc_bg=GRN_L if wc>=3000 else AMB_L if wc>=1500 else RED_L
    wc_fg=GREEN if wc>=3000 else AMB if wc>=1500 else RED
    x=ws5.cell(r,3,f"{wc:,} words");x.fill=F(wc_bg);x.font=Ft(wc_fg,True,10)
    x.alignment=Al("center","center",False);x.border=Br()
    h2_bg=GRN_L if h2>=9 else AMB_L if h2>=6 else RED_L
    h2_fg=GREEN if h2>=9 else AMB if h2>=6 else RED
    x=ws5.cell(r,4,str(h2));x.fill=F(h2_bg);x.font=Ft(h2_fg,True,10)
    x.alignment=Al("center","center",False);x.border=Br()
    TC(ws5,r,5,faq); TC(ws5,r,6,pricing); TC(ws5,r,7,doc)
    cit_bg=GRN_L if cit not in("No","None") else RED_L
    cit_fg=GREEN if cit not in("No","None") else RED
    x=ws5.cell(r,8,cit);x.fill=F(cit_bg);x.font=Ft(cit_fg,True,9)
    x.alignment=Al("center","center",False);x.border=Br()
    C(ws5,r,9,h2_list,bg,TEXT,False,9,"left")

# ── WINNING TEMPLATE NOTE
rh(ws5,14,14)
sec_hdr(ws5,15,1,9,"WINNING PAGE STRUCTURE TEMPLATE  —  Derived from top-ranking hospital & clinic pages (Kokilaben + Apollo + DermaVue + Chaitanya)",NAVY,WHITE,11)
TEMPLATE=[
    ("H1","[Treatment Name] in [City/Region]  or  [Benefit statement] — [Treatment] in [City]","Kokilaben, Oliva, DermaVue all use location + treatment in H1. Google uses H1 as primary relevance signal."),
    ("Intro Para (100-150 words)","Direct answer to what the user just searched — defines treatment, states what conditions it treats, and states Regencare's unique advantage in 2-3 sentences","Scannable intro = lower bounce rate + featured snippet eligibility"),
    ("H2-1 — What is [Treatment]?","Plain English definition (40-60 words) — the exact format that wins voice answers and Google AI Overviews","Manipal, Kokilaben, Apollo Spectra all use this pattern"),
    ("H2-2 — Conditions Treated","Bulleted or card list — 6-10 conditions with brief description each","Stem Cell Care India, Apollo Spectra, RegenOrthoSport"),
    ("H2-3 — Who is Eligible?","Patient criteria — builds qualified traffic (filters non-buyers, warms buyers)","Apollo Spectra, Kokilaben"),
    ("H2-4 — How Does It Work? / Our Process","Step-by-step numbered list — reassures patient, earns HowTo schema","Kokilaben, DermaVue"),
    ("H2-5 — Benefits","4-6 benefits with icons or bullet points — conversion-optimised section","All top performers"),
    ("H2-6 — [Treatment] vs. Alternatives","Comparison table: treatment vs. surgery vs. medication — positions Regencare without making prohibited superiority claims","Kokilaben PRP page — strongest example found"),
    ("H2-7 — Cost / What to Expect","Price range or cost factors — even if exact price not given, cost section targets 'GFC therapy cost Kerala' queries","DermaVue is only local clinic doing this — 170 vol/mo with zero strong answer"),
    ("H2-8 — Why Choose Regencare?","Doctor credentials + branch locations + unique claims (First in South India, Dr. Vineeth MB MS Ortho)","Every top competitor has this section — Regencare's current pages don't"),
    ("H2-9 — FAQs (8-12 questions)","Direct Q&A format, 40-60 words per answer — FAQPage schema applied to each Q&A","Manipal, Kokilaben, Apollo, Chaitanya, Stem Cell Care India all have this"),
    ("H2-10 — Related Treatments / Next Steps","Internal links to related treatment pages — reduces bounce, builds internal authority","Apollo Spectra sidebar links to 6 related treatments"),
    ("CTA Block (repeated 3x)","Book Appointment [Cal.com link] + WhatsApp [MSG91] + Phone — appear at top, middle, and bottom of page","DermaVue: 3 CTAs, Oliva: 4 CTAs, Kokilaben: multiple CTAs"),
    ("Schema Types Required","MedicalClinic + MedicalProcedure + FAQPage + BreadcrumbList + Person (doctor) + LocalBusiness (per branch)","Apollo Spectra: Breadcrumb + MedicalProcedure. Kokilaben: implied full schema. DermaVue: 4 schema types"),
    ("Doctor Authorship","Named doctor (Dr. Vineeth MB, MBBS MS Ortho) in author byline + structured data Person schema — links to /doctors/dr-vineeth-mb","Manipal: best practice — named MD with credentials on every clinical page"),
    ("Academic Citation (1-2 per page)","Link to 1-2 PubMed or peer-reviewed studies supporting the treatment claim","Chaitanya: 6 citations — highest E-E-A-T signal. Manipal: 1 PubMed link"),
]
for i,(element,spec,source) in enumerate(TEMPLATE):
    r=16+i; bg=BLU_L if i%2==0 else WHITE; rh(ws5,r,44)
    C(ws5,r,1,element,bg,BLUE,True,10,"left")
    ws5.merge_cells(start_row=r,start_column=2,end_row=r,end_column=6)
    C(ws5,r,2,spec,bg,TEXT,False,9,"left")
    ws5.merge_cells(start_row=r,start_column=7,end_row=r,end_column=9)
    C(ws5,r,7,f"Source: {source}",bg,GREY,False,9,"left")

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 6: SEO STRATEGY MATRIX
# ═══════════════════════════════════════════════════════════════════════════
ws6=wb.create_sheet("SEO STRATEGY MATRIX")
ws6.sheet_view.showGridLines=False
cw(ws6,1,30)
for c in range(2,12): cw(ws6,c,13)
cw(ws6,12,13); cw(ws6,13,13)

section_title(ws6,1,1,13,"SEO STRATEGY MATRIX  —  Which Strategies Each Competitor Uses  (and Regencare's Current Status)")
sub_title(ws6,2,1,13,"✓ = Implemented fully  ~ = Partial  ✗ = Missing  — ratings based on direct page audit and SERP analysis")

STRATS=[
    "Single clean H1 per page",
    "2,500+ word treatment pages",
    "FAQPage schema",
    "MedicalProcedure schema",
    "BreadcrumbList schema",
    "LocalBusiness schema per branch",
    "Person schema (doctor)",
    "Named doctor with credentials on page",
    "Academic citations (PubMed)",
    "Internal hub-and-spoke linking",
    "Pricing / cost page",
    "Real online booking (not form)",
    "Treatment comparison table",
    "Patient review integration",
    "Location-specific pages",
    "Blog / research content section",
    "Video content embedded",
    "Google Business Profile optimised",
    "Breadcrumb navigation visible",
    "Mobile-first page speed",
]

ALL_SITES=[
    ("REGENCARE.IN",["No","No","No","Partial","No","Partial","Partial","Yes","No","No","Partial","No","No","Partial","Partial","No","No","Partial","No","Yes"]),
    ("Apollo Hospitals",["Yes","Yes","Yes","Yes","Yes","Yes","Yes","Yes","No","Yes","No","Yes","No","Yes","Yes","Yes","No","Yes","Yes","Yes"]),
    ("Apollo Spectra",["Yes","Partial","Yes","Yes","Yes","Yes","Yes","Yes","No","Yes","No","Yes","No","Yes","Yes","Yes","No","Yes","Yes","Yes"]),
    ("Manipal Hospitals",["Yes","Partial","Yes","Yes","Yes","Yes","Yes","Yes","Yes","Yes","No","Yes","No","Yes","Yes","Yes","Yes","Yes","Yes","Yes"]),
    ("Kokilaben Hospital",["Yes","Yes","Yes","Yes","Yes","Yes","Yes","Yes","No","Yes","No","Yes","Yes","Yes","Partial","Yes","No","Yes","Yes","Yes"]),
    ("Stem Cell Care India",["Yes","Yes","Yes","Partial","Partial","Yes","Partial","Partial","No","Yes","No","Yes","No","Yes","Yes","Yes","No","Yes","Yes","Yes"]),
    ("RegenOrthoSport",["Yes","Partial","Yes","Partial","Partial","Yes","Yes","Yes","No","Yes","No","Yes","No","Yes","Yes","Yes","Yes","Yes","Partial","Yes"]),
    ("Chaitanya Stem Cell",["Yes","Yes","Yes","Partial","Partial","Yes","Yes","Yes","Yes","Yes","No","Yes","No","Yes","Partial","Yes","No","Partial","Partial","Yes"]),
    ("DermaVue",["Yes","Yes","No","Partial","Yes","Yes","Partial","Yes","Yes","Yes","Yes","Yes","No","Yes","Yes","Yes","No","Yes","Yes","Yes"]),
    ("Oliva Clinic",["Yes","Yes","Partial","Yes","Yes","Yes","Yes","Yes","No","Yes","Yes","Yes","No","Yes","Yes","Yes","Yes","Yes","Yes","Yes"]),
    ("Orthogen Care",["Yes","Partial","No","No","No","No","Yes","Yes","No","Partial","No","Yes","No","Partial","No","Yes","No","Partial","No","Yes"]),
    ("Cutis International",["Partial","Partial","No","No","No","Yes","Yes","Yes","No","Yes","No","Yes","No","Yes","Yes","No","No","Yes","No","Yes"]),
]

H(ws6,3,1,"SEO Strategy / Signal",NAVY,WHITE,True,10,"left")
for ci,n in enumerate([s[0] for s in ALL_SITES],2):
    fg=WHITE if n!="REGENCARE.IN" else WHITE
    bg=TEAL if n=="REGENCARE.IN" else NAVY
    H(ws6,3,ci,n,bg,fg,True,8)
rh(ws6,3,44)

for si,strat in enumerate(STRATS):
    r=4+si; bg=LIGHT if si%2==0 else WHITE; rh(ws6,r,24)
    C(ws6,r,1,strat,bg,NAVY,True,10,"left")
    for ci,(site,values) in enumerate(ALL_SITES,2):
        TC(ws6,r,ci,values[si])

ws6.freeze_panes="B4"

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 7: ON-PAGE SEO AUDIT
# ═══════════════════════════════════════════════════════════════════════════
ws7=wb.create_sheet("ON-PAGE SEO AUDIT")
ws7.sheet_view.showGridLines=False
ACOLS=["Title Tag Opt.","Single H1","Meta Desc.","Schema Types","FAQ Page","Word Count","Pricing Listed","Real Booking","NMC Compliant","Core Web Vitals"]
cw(ws7,1,28)
for i,w in enumerate([13,11,13,20,10,12,13,13,13,13],2): cw(ws7,i,w)

section_title(ws7,1,1,11,"ON-PAGE SEO AUDIT  —  Local Clinics + Major Hospitals + Regencare")
sub_title(ws7,2,1,11,"Audited via direct page fetch · Strong/Partial/None ratings based on actual page content found")

H(ws7,3,1,"Competitor",NAVY,WHITE,True,10,"left")
for ci,h in enumerate(ACOLS,2): H(ws7,3,ci,h,NAVY,WHITE,True,9)
rh(ws7,3,40)

AUDIT_ROWS=[
    ("REGENCARE.IN (Client)","Partial","No — 7 H1 tags","Partial","Partial (no MedProc)","No",3200,"Partial","No — form only","Yes","Unknown",True),
    ("Apollo Hospitals","Strong","Yes","Strong","Strong (all types)","Yes",1300,"No","Yes","Yes","Strong",False),
    ("Apollo Spectra","Strong","Yes","Strong","Breadcrumb+Medical","Yes",1300,"No","Yes","Yes","Strong",False),
    ("Manipal Hospitals","Strong","Yes","Strong","Strong","Yes",1300,"No","Yes","Yes","Strong",False),
    ("Kokilaben Hospital","Strong","Yes","Strong","Strong","Yes",2100,"No","Yes","Yes","Strong",False),
    ("Stem Cell Care India","Strong","Yes","Strong","Partial","Yes",8000,"No","Yes","Yes","Medium",False),
    ("RegenOrthoSport","Strong","Yes","Strong","Partial","Yes",1800,"No","Yes","Yes","Unknown",False),
    ("Chaitanya Stem Cell","Strong","Yes","Partial","Partial","Yes",3500,"No","Yes","Yes","Unknown",False),
    ("orthogencare.com","Strong","Yes","Partial","None","No",2000,"No","Yes","Yes","Unknown",False),
    ("dermavue.com","Strong","Yes","Strong","Strong (4 types)","No",4900,"Yes","Yes","Yes","Good",False),
    ("olivaclinic.com","Strong","Yes","Strong","Strong","Yes",3200,"Yes","Yes","Yes","Good",False),
    ("cutisinternational.com","Partial","Weak","Partial","Partial","No",2100,"No","Yes","Partial","Unknown",False),
    ("ladensitae.com","Partial","Yes","Partial","None","No",1200,"Partial","Yes","Partial","Unknown",False),
    ("zaayaskinclinic.com","Strong","Yes","Partial","None","No",2300,"No","Yes","Yes","Unknown",False),
]
for i,(name,title,h1,meta,schema,faq,wc,pricing,booking,nmc,cwv,is_regen) in enumerate(AUDIT_ROWS):
    r=4+i; bg=TEA_L if is_regen else (LIGHT if i%2==1 else WHITE); rh(ws7,r,32)
    C(ws7,r,1,name,bg,TEAL if is_regen else NAVY,True,10,"left")
    TC(ws7,r,2,title); TC(ws7,r,3,h1); TC(ws7,r,4,meta)
    schema_bg=GRN_L if "Strong" in str(schema) else AMB_L if "Partial" in str(schema) else RED_L
    schema_fg=GREEN if "Strong" in str(schema) else AMB if "Partial" in str(schema) else RED
    x=ws7.cell(r,5,schema);x.fill=F(schema_bg);x.font=Ft(schema_fg,True,9)
    x.alignment=Al("center","center",True);x.border=Br()
    TC(ws7,r,6,faq)
    wc_bg=GRN_L if wc>=3000 else AMB_L if wc>=1500 else RED_L
    wc_fg=GREEN if wc>=3000 else AMB if wc>=1500 else RED
    x=ws7.cell(r,7,f"{wc:,} words");x.fill=F(wc_bg);x.font=Ft(wc_fg,True,10)
    x.alignment=Al("center","center",False);x.border=Br()
    TC(ws7,r,8,pricing); TC(ws7,r,9,booking); TC(ws7,r,10,nmc); TC(ws7,r,11,cwv)

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 8: AUTHORITY & TRUST
# ═══════════════════════════════════════════════════════════════════════════
ws8=wb.create_sheet("AUTHORITY & TRUST")
ws8.sheet_view.showGridLines=False
for i,w in enumerate([28,12,12,11,14,14,36,10,10],1): cw(ws8,i,w)

section_title(ws8,1,1,9,"AUTHORITY & TRUST SIGNALS  —  Local Clinics + Major Hospitals vs. Regencare")
sub_title(ws8,2,1,9,"DA = Domain Authority estimate · E-E-A-T = Experience, Expertise, Authoritativeness, Trustworthiness · Higher = Better")

for ci,h in enumerate(["Competitor","Google ★","Reviews","Est. DA","Backlinks","Accreditation","Doctor Credentials","E-E-A-T","Content Auth."],1):
    H(ws8,3,ci,h,NAVY,WHITE,True,9)
rh(ws8,3,40)

AUTH_ALL=[
    ("REGENCARE.IN (Client)",4.2,150,18,"Low","None visible","Dr. Vineeth MB — MBBS MS Ortho (dedicated profile)","Medium","Medium",True),
    ("Apollo Hospitals",4.7,"10,000s",72,"Very High","NABH, NABL, JCI","Named specialists + full department","Very High","Very High",False),
    ("Apollo Spectra",4.6,"5,000s",65,"Very High","NABH, NABL","Named specialists linked","Very High","Very High",False),
    ("Manipal Hospitals",4.6,"10,000s",70,"Very High","NABH, NABL, ISO","Dr. Lokesh A Veerappa named on page","Very High","Very High",False),
    ("Kokilaben Hospital",4.8,"5,000s",62,"High","NABH, NABL, JCI","Named physicians + 10yr experience","Very High","Very High",False),
    ("Stem Cell Care India",4.5,None,44,"Medium","NABH partner, GMP, ISO","Team reference (not individually named)","High","High",False),
    ("RegenOrthoSport",4.4,None,32,"Medium","IOF Member","4 named MDs with specialties","High","High",False),
    ("Chaitanya Stem Cell",4.5,None,28,"Low","Awards listed","Dr. Anant Bagul — named orthopedician","Medium","Medium",False),
    ("orthogencare.com",None,None,22,"Low","None","Dr. Vineeth MB — MBBS MS Ortho","Medium","Medium",False),
    ("dermavue.com",4.8,1438,34,"Medium","IADVL, ISO","Dr. Minu Liz Mathew — MBBS MD DVL","High","High",False),
    ("olivaclinic.com",4.9,1475,52,"High","US-FDA equip.","4 MD Dermatologists listed (Kochi)","High","Very High",False),
    ("cutisinternational.com",4.7,500,38,"Medium","Claimed","Multiple specialists","Medium","Medium",False),
    ("ladensitae.com",4.5,None,29,"Low","None","No named doctors on page","Low","Low",False),
    ("zaayaskinclinic.com",4.6,None,18,"Low","None","Unnamed TM consultant","Low","Medium",False),
]

for i,(name,rating,reviews,da,bl,accred,doc,eeat,cauth,is_regen) in enumerate(AUTH_ALL):
    r=4+i; bg=TEA_L if is_regen else (LIGHT if i%2==1 else WHITE); rh(ws8,r,36)
    C(ws8,r,1,name,bg,TEAL if is_regen else NAVY,True,10,"left")
    if rating:
        rb=GRN_L if rating>=4.7 else AMB_L if rating>=4.3 else RED_L
        rf=GREEN if rating>=4.7 else AMB if rating>=4.3 else RED
        x=ws8.cell(r,2,f"{rating} ★");x.fill=F(rb);x.font=Ft(rf,True,11)
        x.alignment=Al("center","center",False);x.border=Br()
    else: TC(ws8,r,2,"Unknown")
    if reviews:
        rv=reviews if isinstance(reviews,int) else 999999
        rb=GRN_L if rv>=1000 else AMB_L if rv>=300 else RED_L
        rf=GREEN if rv>=1000 else AMB if rv>=300 else RED
        rv_str=f"{reviews:,}+" if isinstance(reviews,int) else str(reviews)
        x=ws8.cell(r,3,rv_str);x.fill=F(rb);x.font=Ft(rf,True,10)
        x.alignment=Al("center","center",False);x.border=Br()
    else: TC(ws8,r,3,"Unknown")
    da_bg=GRN_L if da>=50 else AMB_L if da>=25 else RED_L
    da_fg=GREEN if da>=50 else AMB if da>=25 else RED
    x=ws8.cell(r,4,str(da));x.fill=F(da_bg);x.font=Ft(da_fg,True,10)
    x.alignment=Al("center","center",False);x.border=Br()
    TC(ws8,r,5,bl)
    accred_bg=GRN_L if "NABH" in str(accred) or "JCI" in str(accred) else AMB_L if "ISO" in str(accred) or "IADVL" in str(accred) else RED_L
    x=ws8.cell(r,6,accred);x.fill=F(accred_bg);x.font=Ft(TEXT,False,9)
    x.alignment=Al("left","center",True);x.border=Br()
    C(ws8,r,7,doc,bg,TEXT,False,9,"left")
    TC(ws8,r,8,eeat); TC(ws8,r,9,cauth)

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 9: SERVICE COVERAGE (simplified)
# ═══════════════════════════════════════════════════════════════════════════
ws9=wb.create_sheet("SERVICE COVERAGE")
ws9.sheet_view.showGridLines=False
cw(ws9,1,32); cw(ws9,2,14)
for c in range(3,15): cw(ws9,c,12)

section_title(ws9,1,1,14,"SERVICE COVERAGE MAP  —  Local Clinics + Major Hospitals vs. Regencare")
sub_title(ws9,2,1,14,"✓ Yes = Dedicated page exists  ~ Partial = Mentioned but no dedicated page  ✗ No = Not offered or no online presence")

SVCS=["GFC Therapy","PRP — Hair","PRP — Ortho/Joints","Stem Cell Therapy","Hair Transplant (FUE)","Exosome Therapy",
      "Sports Injury (Regen)","Knee Pain (non-surgical)","Spine / Back Pain",
      "UAE / Medical Tourism Page","Named Doctor Profile","Blog / Research","Real Online Booking","Cost / Pricing Page"]

REGEN_SVC=["Yes","Yes","Yes","Yes","No","Yes","Yes","Yes","Yes","Partial","Yes","No","No","Partial"]
HOSP_SVC={
    11:["No","No","Yes","Yes","No","No","Yes","Yes","Yes","Partial","Yes","Yes","Yes","No"],
    12:["No","No","Yes","Yes","No","No","Yes","Yes","Yes","Partial","Yes","Yes","Yes","No"],
    13:["No","No","Yes","Yes","No","No","Yes","Yes","Yes","No","Yes","Yes","Yes","No"],
    14:["No","No","Yes","No","No","No","Yes","Yes","Yes","No","Yes","Yes","Yes","No"],
    15:["No","No","Yes","Yes","No","Yes","Yes","Yes","Yes","Yes","Partial","Yes","Yes","No"],
    16:["No","No","Yes","Yes","No","No","Yes","Yes","Yes","Yes","Yes","Yes","Yes","No"],
    17:["No","No","Yes","Yes","No","No","Yes","Yes","Yes","No","Yes","Yes","Yes","No"],
}
LOCAL_SVC={
    1:["No","No","Yes","Partial","No","No","Yes","Yes","Yes","No","Yes","Yes","Yes","No"],
    2:["Yes","Yes","No","No","Yes","No","No","No","No","No","Yes","Yes","Yes","Yes"],
    3:["No","Yes","No","No","Yes","Yes","No","No","No","No","Yes","Yes","Yes","Yes"],
    4:["Yes","Yes","No","No","Yes","Yes","No","No","No","Yes","Yes","No","Yes","No"],
    5:["No","Yes","No","No","Yes","No","No","No","No","Yes","No","Partial","Yes","Partial"],
    6:["Yes","Yes","No","No","No","No","No","No","No","No","No","Yes","Yes","No"],
    7:["No","No","No","No","Yes","No","No","No","No","Yes","No","No","Yes","No"],
    8:["No","No","Yes","Yes","No","No","Yes","Yes","Yes","No","Yes","Partial","Yes","No"],
    9:["Yes","Yes","No","No","No","No","No","No","No","No","No","No","Yes","Partial"],
   10:["No","No","Yes","Yes","No","No","Yes","Yes","Yes","No","No","Partial","Yes","No"],
}

H(ws9,3,1,"Service / Treatment",NAVY,WHITE,True,10,"left")
H(ws9,3,2,"REGENCARE",TEAL,WHITE,True,9)
for ci,(cid,name,*_) in enumerate(HOSPITAL_COMPS,3): H(ws9,3,ci,name[:10],PURP,WHITE,True,9)
for ci,(cid,name,*_) in enumerate(LOCAL_COMPS,10):
    if ci<=14: H(ws9,3,ci,name[:10],NAVY,WHITE,True,9)
rh(ws9,3,40)

for si,svc in enumerate(SVCS):
    r=4+si; bg=LIGHT if si%2==0 else WHITE; rh(ws9,r,24)
    C(ws9,r,1,svc,bg,NAVY,True,10,"left")
    TC(ws9,r,2,REGEN_SVC[si])
    for ci,cid in enumerate(range(11,18),3): TC(ws9,r,ci,HOSP_SVC[cid][si])
    for ci,cid in enumerate(range(1,6),10):
        if ci<=14: TC(ws9,r,ci,LOCAL_SVC[cid][si])

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 10: WHITE SPACE & GAPS
# ═══════════════════════════════════════════════════════════════════════════
ws10=wb.create_sheet("WHITE SPACE & GAPS")
ws10.sheet_view.showGridLines=False
for i,w in enumerate([14,40,30,16,14,52],1): cw(ws10,i,w)

section_title(ws10,1,1,6,"WHITE SPACE & GAP OPPORTUNITIES  —  What No Competitor Owns Well")
sub_title(ws10,2,1,6,"Including gaps vs. both local clinics AND major national hospital sites — Regencare's fastest page-1 paths")

for ci,h in enumerate(["Priority","Gap Opportunity","Why It's Unowned","Est. Keyword Volume","Difficulty","Recommended Action for Regencare"],1):
    H(ws10,3,ci,h,NAVY,WHITE,True,10)
rh(ws10,3,36)

WS=[
    ("CRITICAL","Stem cell therapy Kerala — comprehensive guide",
     "Major hospitals (Apollo/Manipal) rank nationally but have no Kerala-specific stem cell page. Local clinics have nothing. Near-zero local competition.","High — 480+/mo","Easy",
     "Publish 3,500-word 'Stem Cell Therapy Kerala' page: what it is, conditions, eligibility, process, cost, FAQs (FAQ schema). Outranks all local competition immediately, challenges hospitals on location signal."),
    ("CRITICAL","Ortho cannibalization resolution — orthogencare.com",
     "Same doctor on competing domain. No other competitor has this structural problem. Every fix to Regencare's ortho pages is partially undermined until resolved.","All ortho keywords","Immediate",
     "Define clear domain strategy: either 301 redirect orthogencare.com → regencare.in (if Dr. Vineeth MB agrees), OR create strict content differentiation (orthogencare = pure PRP research, regencare = full clinic with booking). Implement within 30 days."),
    ("CRITICAL","Treatment cost / pricing pages for all major treatments",
     "No local clinic (except DermaVue partially) publishes prices. No hospital publishes prices. 'GFC therapy cost Kerala', 'PRP treatment cost Kochi', 'stem cell therapy cost India' — all high-intent queries with no strong answer page.","High — 170–480/mo each","Easy",
     "Create /pricing page + individual treatment cost sections. Even 'starting from ₹X' with cost factors listed captures the transactional intent query cluster entirely."),
    ("CRITICAL","GFC therapy — comprehensive hub (3,000+ words)",
     "DermaVue ranks #1 with 4,900 words. Regencare is at #9 with shorter content and no FAQ schema. Content depth is the ONLY gap — not domain authority.","1,300/mo — GFC Kochi","Easy",
     "Rewrite GFC therapy page to 3,000+ words: What is GFC vs PRP (comparison table), 12+ H2 sections, 10 FAQs with FAQ schema, doctor quote, cost section, before/after timeline. Target: #1-3 within 90 days of publish."),
    ("HIGH","UAE / NRI medical tourism landing page",
     "Cutis has UAE clinics but no Kerala NRI-facing regen content. Apollo has international pages but no Kerala-specific. Zero competition for 'regenerative medicine India for NRI', 'Kerala ortho clinic for UAE residents'.","Medium — 18 keywords","Easy",
     "Build /medical-tourism/nri page: location (Kochi, Calicut, Chennai), treatments available, Dr. Vineeth MB credentials, how to book from UAE, accommodation guidance, WhatsApp CTA."),
    ("HIGH","Dr. Vineeth MB comprehensive authority hub",
     "No competitor (including orthogencare) has a comprehensive doctor profile with original Q&A, research commentary, and media. This is Regencare's only uncopied competitive moat.","Medium — all branded queries","Easy",
     "Build /doctors/dr-vineeth-mb: full credentials, interview-style Q&A, original clinical commentary, published works, speaking, media mentions. Person schema. Link from every treatment page as author."),
    ("HIGH","Knee pain non-surgical treatment — regen approach (vs. Ayurveda)",
     "Ayurveda clinics dominate 'knee pain treatment without surgery Kerala'. No modern regen clinic has claimed this space. Regencare already ranks #2 — a content upgrade would secure #1.","Medium — 720/mo","Easy",
     "Expand knee pain page: add comparison table (PRP vs Ayurveda vs Surgery vs Physiotherapy), add condition stages (mild/moderate/severe OA), patient eligibility criteria, FAQ schema (10 Q&As). 2,500+ words."),
    ("HIGH","Calicut / Kozhikode treatment-specific pages",
     "All competitors focus content on Kochi. Calicut has 30-40% lower keyword competition for equivalent queries with no quality local answer pages. GBP local pack in Calicut also under-served.","Medium — 260-390/mo Calicut","Easy",
     "Mirror Kochi treatment pages for Calicut: /gfc-therapy/calicut, /prp-treatment/calicut, /stem-cell-therapy/calicut — same content framework, location-specific text. GBP Calicut optimisation in parallel."),
    ("HIGH","Sports injury treatment — regen medicine Kerala",
     "Physiotherapy and Ayurveda own 'sports injury treatment Kerala'. No regen clinic has a dedicated sports injury page targeting ACL, PCL, meniscal tear, tennis elbow — all treatable with PRP/stem cell.","Medium — 300+/mo","Medium",
     "Build /conditions/sports-injuries: list 8-10 conditions, explain PRP/stem cell approach for each, eligibility, timeline, FAQ schema. Internal link from all treatment pages."),
    ("MEDIUM","GFC therapy vs. PRP vs. Stem Cell comparison page",
     "No competitor — local or national — has a comparison page for these three treatments. Users searching 'GFC vs PRP' find no authoritative answer. Featured snippet opportunity.","Medium — 170+/mo","Easy",
     "Build /treatments/gfc-vs-prp-vs-stem-cell: side-by-side table, condition suitability chart, cost comparison, which one Regencare recommends and why. Targets comparison queries + captures commercial intent."),
    ("MEDIUM","Chennai regenerative medicine — market entry content",
     "No competitor has strong Chennai-specific regen content. Chennai is the newest Regencare branch and has the most online competition — but also the most search volume.","Medium — 480+/mo Chennai","Medium",
     "Build Chennai treatment pages: /branches/chennai, /gfc-therapy/chennai, /prp-treatment/chennai. Optimise Chennai GBP. NAP consistency fix first (email mismatch)."),
    ("MEDIUM","Spine / back pain non-surgical — regen approach",
     "Physio and ortho hospitals own spine content. No regen clinic has a spine/back pain page specifically. PRP for disc herniation, prolotherapy for chronic back pain — untapped.","Medium","Medium",
     "Build /conditions/back-pain-spine: disc herniation, lumbar pain, sciatica — PRP/regen approach. Internal link from ortho and sports injury pages."),
    ("MEDIUM","Research / outcomes blog — Regencare clinical commentary",
     "Manipal and Chaitanya use academic citations as E-E-A-T signals. No Kerala clinic publishes original commentary on regen research. Dr. Vineeth MB's commentary on a PubMed paper = E-E-A-T gold.","Low vol, High authority","Medium",
     "Dr. Vineeth MB writes 4-6 research commentary articles/year. Reference PubMed studies. These build domain-wide E-E-A-T and earn natural backlinks from medical directories and health journalists."),
    ("LOW","Google AI Overview + ChatGPT citation strategy (GEO)",
     "No Kerala regen clinic is being cited by AI engines. Apollo and Manipal are cited nationally. Publishing structured answer content trains LLMs to cite Regencare for Kerala-specific queries.","AI traffic — growing","Hard",
     "Publish citation-bait content: 'What is the best regenerative medicine clinic in Kerala?' structured as factual entity — consistent NAP, Wikipedia-style About page, Google Knowledge Panel claim, FAQ schema on every page."),
    ("LOW","Video SERPs — YouTube SEO for treatment keywords",
     "Video carousels appear for GFC, PRP, and knee pain keywords. No Kerala regen clinic has YouTube presence. Manipal and Apollo have videos but not Kochi-location-specific.","Video SERP gap","Medium",
     "Film Dr. Vineeth MB procedure walkthroughs: GFC session, PRP process, knee stem cell. YouTube SEO: keyword titles, chapters, location tags. Embed on treatment pages for dwell time."),
]
prio_c={"CRITICAL":(RED_L,RED),"HIGH":(AMB_L,AMB),"MEDIUM":(TEA_L,TEAL),"LOW":(GRY_L,GREY)}
for i,(p,topic,why,vol,diff,action) in enumerate(WS):
    r=4+i; bg=LIGHT if i%2==0 else WHITE; rh(ws10,r,52)
    pb,pf=prio_c.get(p,(WHITE,TEXT))
    x=ws10.cell(r,1,p);x.fill=F(pb);x.font=Ft(pf,True,9)
    x.alignment=Al("center","center",False);x.border=Br()
    C(ws10,r,2,topic,bg,NAVY,True,10,"left")
    C(ws10,r,3,why,bg,TEXT,False,9,"left")
    C(ws10,r,4,vol,bg,TEAL,True,9,"center")
    d_bg=GRN_L if diff=="Easy" else AMB_L if diff=="Medium" else RED_L
    d_fg=GREEN if diff=="Easy" else AMB if diff=="Medium" else RED
    x=ws10.cell(r,5,diff);x.fill=F(d_bg);x.font=Ft(d_fg,True,9)
    x.alignment=Al("center","center",False);x.border=Br()
    C(ws10,r,6,action,bg,TEXT,False,9,"left")

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 11: REGENCARE SEO BLUEPRINT
# ═══════════════════════════════════════════════════════════════════════════
ws11=wb.create_sheet("REGENCARE SEO BLUEPRINT")
ws11.sheet_view.showGridLines=False
for i,w in enumerate([22,10,10,10,44,28,22],1): cw(ws11,i,w)

section_title(ws11,1,1,7,"REGENCARE.IN  —  COMPLETE SEO BLUEPRINT FOR PAGE-1 RANKINGS")
sub_title(ws11,2,1,7,"Derived from analysis of Apollo, Manipal, Kokilaben, DermaVue + 10 local competitors — every item is evidence-backed")

# ─── SECTION A: URL & SITE ARCHITECTURE ────────────────────────────────────
rh(ws11,3,14)
sec_hdr(ws11,4,1,7,"A.  SITE ARCHITECTURE  —  Hub-and-Spoke URL Structure (Derived from Apollo Spectra + Manipal model)",NAVY,WHITE,11)

H(ws11,5,1,"URL Pattern",NAVY,WHITE,True,9,"left")
H(ws11,5,2,"Page Type",NAVY,WHITE,True,9)
H(ws11,5,3,"Priority",NAVY,WHITE,True,9)
H(ws11,5,4,"Schema Required",NAVY,WHITE,True,9)
H(ws11,5,5,"Target Keyword Cluster",NAVY,WHITE,True,9,"left")
H(ws11,5,6,"Competitive Model",NAVY,WHITE,True,9,"left")
H(ws11,5,7,"Content Length",NAVY,WHITE,True,9)
rh(ws11,5,32)

ARCH=[
    ("/","Homepage","CRITICAL","MedicalClinic + LocalBusiness + Organization","regenerative medicine Kerala / Regencare","Apollo Hospitals homepage structure","2,000+ words"),
    ("/treatments/gfc-therapy/","Treatment Hub","CRITICAL","MedicalProcedure + FAQPage","GFC therapy Kerala / GFC therapy Kochi","DermaVue 4,900-word GFC page","3,000+ words"),
    ("/treatments/gfc-therapy/kochi/","Treatment + Location","HIGH","MedicalProcedure + LocalBusiness","GFC therapy Kochi (1,300/mo)","Apollo Spectra location pages","2,000+ words"),
    ("/treatments/gfc-therapy/calicut/","Treatment + Location","HIGH","MedicalProcedure + LocalBusiness","GFC therapy Calicut (260/mo)","Apollo Spectra location pages","1,800+ words"),
    ("/treatments/prp-therapy/","Treatment Hub","HIGH","MedicalProcedure + FAQPage","PRP treatment Kerala (1,600/mo)","Kokilaben PRP page — 9 H2 model","2,500+ words"),
    ("/treatments/stem-cell-therapy/","Treatment Hub","HIGH","MedicalTherapy + FAQPage","stem cell therapy Kerala (480/mo)","Chaitanya + Stem Cell Care India","3,500+ words"),
    ("/treatments/stem-cell-therapy/kerala/","Treatment + Location","HIGH","MedicalTherapy + LocalBusiness","stem cell therapy Kerala — local","Manipal + Apollo Spectra pattern","2,000+ words"),
    ("/conditions/knee-pain/","Condition Hub","HIGH","MedicalCondition + FAQPage","knee pain treatment without surgery (720/mo)","RegenOrthoSport 12-H2 model","2,500+ words"),
    ("/conditions/hair-loss/","Condition Hub","HIGH","MedicalCondition + FAQPage","hair loss treatment Kochi (2,400/mo)","Oliva Clinic: Cost+Procedure+Results+Reviews","2,500+ words"),
    ("/conditions/sports-injuries/","Condition Hub","MEDIUM","MedicalCondition + FAQPage","sports injury treatment Kerala","RegenOrthoSport structure","2,000+ words"),
    ("/conditions/spine-back-pain/","Condition Hub","MEDIUM","MedicalCondition + FAQPage","back pain non-surgical Kerala","Gap — no competitor owns it","2,000+ words"),
    ("/doctors/dr-vineeth-mb/","Doctor Profile","HIGH","Person + MedicalBusiness","Dr. Vineeth MB regenerative medicine","Manipal named-doctor model","1,500+ words"),
    ("/branches/kochi/","Branch Page","HIGH","LocalBusiness + MedicalClinic","regenerative medicine Kochi (170/mo)","Apollo location page pattern","1,200+ words"),
    ("/branches/calicut/","Branch Page","HIGH","LocalBusiness + MedicalClinic","regenerative medicine Calicut (260/mo)","Apollo location page pattern","1,200+ words"),
    ("/branches/chennai/","Branch Page","HIGH","LocalBusiness + MedicalClinic","regenerative medicine Chennai (480/mo)","Apollo location page pattern","1,200+ words"),
    ("/medical-tourism/","NRI Hub","MEDIUM","MedicalClinic + FAQPage","regenerative medicine India for NRI","Cutis International UAE positioning","2,000+ words"),
    ("/treatments/compare/","Comparison Page","MEDIUM","FAQPage","GFC vs PRP vs Stem Cell","Gap — no competitor owns it","1,500+ words"),
    ("/pricing/","Pricing Page","HIGH","FAQPage","GFC therapy cost Kerala / PRP cost Kochi","Gap — no competitor owns it","1,000+ words"),
    ("/blog/","Blog Hub","HIGH","Blog + Article + Person","Informational keywords — long tail","Manipal + Apollo blog structure","N/A — ongoing"),
    ("/research/","Research Hub","MEDIUM","ScholarlyArticle + Person","clinical evidence / regen medicine research","Manipal PubMed citation model","N/A — ongoing"),
]
for i,(url,ptype,prio,schema,kw,model,length) in enumerate(ARCH):
    r=6+i; bg=BLU_L if i%2==0 else WHITE; rh(ws11,r,28)
    C(ws11,r,1,url,bg,BLUE,True,9,"left")
    C(ws11,r,2,ptype,bg,TEXT,False,9,"center")
    prio_cell(ws11,r,3,prio)
    C(ws11,r,4,schema,bg,TEAL,False,9,"left")
    C(ws11,r,5,kw,bg,TEXT,False,9,"left")
    C(ws11,r,6,model,bg,GREY,False,9,"left")
    wl_bg=GRN_L if "3,000" in length or "2,500" in length else AMB_L if "2,000" in length or "1,800" in length else RED_L if "N/A" not in length else WHITE
    wl_fg=GREEN if "3,000" in length or "2,500" in length else AMB if "2,000" in length else TEXT
    x=ws11.cell(r,7,length);x.fill=F(wl_bg);x.font=Ft(wl_fg,True,9)
    x.alignment=Al("center","center",False);x.border=Br()

# ─── SECTION B: ON-PAGE TEMPLATE ────────────────────────────────────────────
base=6+len(ARCH)+2
rh(ws11,base-1,14)
sec_hdr(ws11,base,1,7,"B.  ON-PAGE TEMPLATE  —  Every Treatment Page Must Follow This Structure (Derived from Kokilaben + Apollo Spectra + DermaVue)",NAVY,WHITE,11)
H(ws11,base+1,1,"Page Element",NAVY,WHITE,True,9,"left")
H(ws11,base+1,2,"Spec",NAVY,WHITE,True,9,"center")
H(ws11,base+1,3,"Schema",NAVY,WHITE,True,9,"center")
ws11.merge_cells(start_row=base+1,start_column=4,end_row=base+1,end_column=5)
H(ws11,base+1,4,"Content Spec",NAVY,WHITE,True,9,"left")
ws11.merge_cells(start_row=base+1,start_column=6,end_row=base+1,end_column=7)
H(ws11,base+1,6,"Evidence — Which Competitor Does This Best",NAVY,WHITE,True,9,"left")
rh(ws11,base+1,32)

TEMPLATE=[
    ("Title Tag","[Treatment] in [City] | Regencare","None","60 chars max. Include primary keyword first. City if location page.","Kokilaben, Oliva, Apollo Spectra — all lead with treatment + location"),
    ("Meta Description","160 chars. Include: treatment, location, doctor name, CTA.","None","Trigger SERP clicks — include 'Dr. Vineeth MB' and 'Book Free Consultation'.","All major hospital sites — conversion-optimised meta descriptions"),
    ("H1 (one only)","[Treatment] in [City] — [benefit or qualifier]","None","ONE H1. Contains primary keyword. Location included. Current Regencare has 7 H1s — fix immediately.","Every top competitor: single H1. DermaVue: 'PRP & GFC Hair Treatment Near Kochi — 70x More Growth Factors'"),
    ("Intro Para (100-150w)","Direct answer to search query. Define treatment. State conditions. State Regencare's unique advantage.","Speakable","Scannable, 2-3 sentences max per para. Targets featured snippet + voice answer.","Kokilaben, Apollo Spectra — intro directly answers the query in plain English"),
    ("H2 — What is [Treatment]?","40-60 word direct answer. Plain English. No jargon first.","FAQPage Q&A","This exact format wins Google AI Overview citations and voice answers.","Manipal: 'What are Stem Cells?' — plain definition, directly answers query"),
    ("H2 — Conditions Treated","6-10 conditions as cards or bullets. Brief description each.","MedicalIndication","Covers condition-level queries under one treatment page.","Stem Cell Care India: 50+ conditions listed. Apollo Spectra: eligibility criteria section"),
    ("H2 — How Does It Work?","Numbered step-by-step process. 5-7 steps.","HowTo","HowTo schema eligible. Reassures patient. Reduces anxiety-driven bounce.","Kokilaben: numbered procedure steps. DermaVue: session-by-session guide"),
    ("H2 — Benefits","4-6 benefits. Short, scannable. Icons or bullets.","None","Conversion-optimised section. Not a list of features — patient outcomes.","All top performers have this section"),
    ("H2 — [Treatment] vs. Alternatives","Comparison table: treatment vs surgery vs medication vs other.","None","Positions Regencare as superior without making prohibited NMC superiority claims.","Kokilaben PRP page: PRP vs Steroid vs Surgery vs Hyaluronic Acid — strongest example found"),
    ("H2 — Cost / What to Expect","Price range or factors. 'Starting from ₹X' acceptable.","FAQPage","Captures 'GFC therapy cost Kerala' (170/mo) — currently no strong answer on SERPs.","DermaVue: only local competitor listing pricing. Gap for all others including hospitals."),
    ("H2 — Why Regencare?","Dr. Vineeth MB credentials + branch locations + unique claims.","None","Differentiates from all 10 local clinics. Must include: 'South India's first', MS Ortho credentials, multi-branch.","Every hospital has this section. Every local clinic lacks it. Regencare's current pages don't have it."),
    ("H2 — FAQs (10-12 Q&A)","Direct question-answer format. 40-60 words per answer.","FAQPage","10 Q&As = People Also Ask box eligibility for all 10 questions. Each answer is a featured snippet opportunity.","Manipal, Kokilaben, Apollo, Chaitanya, Stem Cell Care India — all use FAQPage schema"),
    ("CTA Block (×3)","Book Appointment [Cal.com] + WhatsApp [MSG91] + Phone","None","Appear at: top of page, after benefits section, after FAQs. Never let user scroll without seeing a CTA.","DermaVue 3 CTAs, Oliva 4 CTAs, Kokilaben 5 CTAs — multiple conversion touchpoints"),
    ("Doctor Authorship","Named: 'Written/Reviewed by Dr. Vineeth MB, MBBS MS Ortho'","Person","Links to /doctors/dr-vineeth-mb. JSON-LD Person schema on every clinical page.","Manipal: gold standard — Dr. Lokesh A Veerappa named with credentials on every blog"),
    ("Academic Citation","1-2 PubMed/peer-reviewed citations per treatment page.","None","Builds E-E-A-T. Shows evidence basis. Required for Google's 'Your Money or Your Life' (YMYL) content.","Chaitanya: 6 PubMed citations — highest found. Manipal: 1 PubMed link on blog"),
    ("Internal Links (5+)","Link to: related treatments, conditions, doctor profile, branch pages, blog posts.","None","Passes authority to new pages. Reduces bounce. Hub-spoke architecture.","Apollo Spectra: 6 related treatment links in sidebar. Manipal: 4 internal links per blog"),
    ("Breadcrumb Navigation","Home > Treatments > GFC Therapy > Kochi","BreadcrumbList","Appears in SERP as rich result. Helps Google understand page hierarchy.","Apollo Spectra: clear breadcrumb on every page. Improves CTR by 15-20%."),
    ("Schema Block (JSON-LD)","MedicalProcedure + FAQPage + Person + LocalBusiness + Breadcrumb","All above","Single JSON-LD block in <head>. Links all entities via @id. Google recommended format.","Apollo: full schema graph. DermaVue: 4 schema types. Regencare currently: partial only"),
]
for i,(elem,spec,schema,content,evidence) in enumerate(TEMPLATE):
    r=base+2+i; bg=BLU_L if i%2==0 else WHITE; rh(ws11,r,44)
    C(ws11,r,1,elem,bg,BLUE,True,10,"left")
    C(ws11,r,2,spec,bg,TEXT,False,9,"center")
    sc_bg=GRN_L if schema not in("None","") else GRY_L
    x=ws11.cell(r,3,schema);x.fill=F(sc_bg);x.font=Ft(GREEN if schema!="None" else GREY,True,9)
    x.alignment=Al("center","center",True);x.border=Br()
    ws11.merge_cells(start_row=r,start_column=4,end_row=r,end_column=5)
    C(ws11,r,4,content,bg,TEXT,False,9,"left")
    ws11.merge_cells(start_row=r,start_column=6,end_row=r,end_column=7)
    C(ws11,r,6,evidence,bg,GREY,False,9,"left")

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 12: ACTION PRIORITIES
# ═══════════════════════════════════════════════════════════════════════════
ws12=wb.create_sheet("ACTION PRIORITIES")
ws12.sheet_view.showGridLines=False
for i,w in enumerate([4,12,36,42,18,10,40],1): cw(ws12,i,w)

section_title(ws12,1,1,7,"COMPETITIVE STUDY ACTION PRIORITIES  —  v2.0  (Local + Hospital Intelligence)")
sub_title(ws12,2,1,7,"Ranked by impact × urgency. Every action is directly supported by evidence from the 17 competitor sites studied.")

for ci,h in enumerate(["#","Priority","Action","Competitive Evidence","Owner","GTM Step","Expected Outcome"],1):
    H(ws12,3,ci,h,NAVY,WHITE,True,10)
rh(ws12,3,36)

ACTIONS=[
    (1,"CRITICAL","Resolve orthogencare.com brand cannibalization",
     "Same doctor appears on identical SERPs. Every other fix is partially undermined until this is resolved. No other competitor has this structural problem.",
     "Strategy + Dev","Step 2A","Stop splitting SERP authority for all ortho/regen queries between two domains owned by the same doctor"),
    (2,"CRITICAL","Fix 7 H1 tags on homepage → single H1",
     "All 17 competitors studied have a single, keyword-containing H1. Apollo, Manipal, Kokilaben, DermaVue — every site studied. Regencare is the only exception.",
     "Dev","Step 2A","Primary keyword signal restored. Estimated ranking improvement within 2-3 crawl cycles (2-4 weeks)"),
    (3,"CRITICAL","Restore Blog (404) + Resources (404)",
     "Manipal, Apollo, Oliva, DermaVue all use blog content to rank for informational keywords. Regencare's blog is entirely absent — the primary long-tail traffic surface is missing.",
     "Dev + Content","Step 2A","Primary content authority surface restored — prerequisite for all content production in Steps 3A/3B"),
    (4,"CRITICAL","Replace contact form with Cal.com real booking",
     "Every major hospital (Apollo, Kokilaben) and top clinic (DermaVue, Oliva) has real appointment booking. Regencare's Book button leads to a form. Zero conversion tracking.",
     "Dev","Step 3A","Direct bookings tracked and confirmed. Estimated 40-60% improvement in conversion rate from organic traffic"),
    (5,"CRITICAL","Fix Chennai NAP inconsistency",
     "NAP consistency is a local pack ranking factor. All studied competitors have clean NAP. Chennai email domain mismatch identified in baseline audit conflicts with GBP.",
     "Marketing","Step 2A","Chennai local pack eligibility restored. GBP and citations aligned."),
    (6,"CRITICAL","Rewrite GFC Therapy page to 3,000+ words with 10 H2s + FAQ schema",
     "DermaVue ranks #1 with 4,900-word page. Regencare is #9 with shorter content. Content depth is the ONLY gap — not DA, not backlinks. DermaVue has no FAQ schema — Regencare can leapfrog with FAQ schema.",
     "Content","Step 3B","Move #9 → Top 3 for GFC Kochi (1,300/mo). Capture 5+ People Also Ask boxes. Estimated 90 days to page 1 position."),
    (7,"HIGH","Implement full schema graph on every page",
     "Apollo, Manipal, Kokilaben all use MedicalProcedure + FAQPage + BreadcrumbList + Person schema as standard. DermaVue uses 4 schema types and outranks Regencare on every shared keyword.",
     "Dev","Step 3A","SERP rich result eligibility: FAQ rich results, breadcrumb display, local pack enhancement, voice answer candidacy"),
    (8,"HIGH","Publish stem cell therapy Kerala comprehensive guide (3,500 words)",
     "Apollo/Manipal rank nationally but have no Kerala-specific stem cell page. Local clinics have nothing. Chaitanya's 3,500-word guide with 6 academic citations is the model to beat.",
     "Content","Step 3B","#1-3 for 'stem cell therapy Kerala' (480/mo) — no quality local competitor. Challenge national hospital sites with location signal."),
    (9,"HIGH","Publish transparent pricing pages for all major treatments",
     "No competitor — local or national — lists prices online. 'GFC therapy cost Kerala' (170/mo), 'stem cell therapy cost India' (480/mo) — zero strong answers on SERPs.",
     "Content","Step 3B","Capture entire cost-query cluster. Highest-converting traffic type (transactional intent). Zero competition currently."),
    (10,"HIGH","Build Dr. Vineeth MB comprehensive authority hub",
     "Manipal's gold standard: named doctor with credentials on every clinical page. Apollo links every treatment page to specialist profiles. Regencare has a basic profile — needs full hub with Q&A, research, media.",
     "Content + SEO","Step 3A","Domain-wide E-E-A-T signal improvement. Every treatment page ranking improves when authored by a credentialed, schema-marked doctor profile."),
    (11,"HIGH","Build UAE/NRI medical tourism landing page",
     "Cutis International has UAE clinics. Apollo has international pages. But no Kerala-based regen clinic has an NRI-facing regen page. 18 keywords, near-zero local competition.",
     "Content","Step 3A","Capture UAE segment — highest patient LTV. No current competitor. Positions Regencare ahead of Cutis for NRI regen queries."),
    (12,"HIGH","Optimise Google Business Profile for all 3 branches",
     "DermaVue (7-clinic GBP network) and Oliva dominate local packs. Kokilaben's GBP is fully optimised. Regencare GBPs have incomplete service listings, low review counts, NAP issues.",
     "Marketing","Step 4A","Local pack presence for Kochi, Calicut, Chennai. Drives call and walk-in conversions directly."),
    (13,"HIGH","Build Calicut-specific treatment pages",
     "All 10 local competitors focus content on Kochi. Calicut has 30-40% lower keyword competition. Apollo Spectra's location-page model (separate page per city × treatment) is the playbook.",
     "Content","Step 3B","Page-1 rankings for Calicut keywords with 40% less content investment than equivalent Kochi pages. GBP local pack in Calicut also improves."),
    (14,"HIGH","Build knee pain non-surgical treatment page (2,500+ words)",
     "Regencare already ranks #2 — a content upgrade secures #1. Ayurveda clinics dominate with thin pages. RegenOrthoSport 12-H2 model shows the content depth needed. Add PRP vs Ayurveda comparison table.",
     "Content","Step 3B","Secure #1 for 'knee pain treatment without surgery Kerala' (720/mo). Treatment comparison table adds commercial intent capture."),
    (15,"MEDIUM","Build sports injury treatment hub",
     "RegenOrthoSport's 12-H2 model covers ACL, PCL, meniscal, tennis elbow. No Kerala clinic has this. Physiotherapy owns the space with thin content.",
     "Content","Step 3B","Capture sports injury segment — high commercial intent, no regen competitor in Kerala"),
    (16,"MEDIUM","Add academic citations to all treatment pages (1-2 PubMed per page)",
     "Chaitanya uses 6 PubMed citations — highest E-E-A-T signal found. Manipal links to PubMed. These signal to Google that content meets YMYL (Your Money or Your Life) quality standards.",
     "Content","Step 3B","E-E-A-T improvement across all treatment pages. Particularly important for Google's medical content quality assessment."),
    (17,"LOW","Begin AI citation / GEO strategy (ChatGPT, Gemini, Perplexity)",
     "No Kerala clinic is cited by AI engines. Apollo and Manipal are cited nationally. Publishing structured answer content + consistent entity data trains LLMs to cite Regencare for Kerala queries.",
     "Content + SEO","Step 5C","Regencare cited in AI Overviews, ChatGPT, and Perplexity for regenerative medicine Kerala queries — growing traffic surface"),
]
for i,(num,prio,action,evidence,owner,gtm,outcome) in enumerate(ACTIONS):
    r=4+i; bg=LIGHT if i%2==0 else WHITE; rh(ws12,r,52)
    C(ws12,r,1,str(num),bg,NAVY,True,11,"center")
    prio_cell(ws12,r,2,prio)
    C(ws12,r,3,action,bg,NAVY,True,10,"left")
    C(ws12,r,4,evidence,bg,TEXT,False,9,"left")
    C(ws12,r,5,owner,bg,TEAL,True,9,"center")
    C(ws12,r,6,gtm,bg,NAVY,True,9,"center")
    C(ws12,r,7,outcome,bg,GREEN,False,9,"left")

# Footer
fr=4+len(ACTIONS)+1
ws12.merge_cells(f"A{fr}:G{fr}")
x=ws12.cell(fr,1,"Prepared by NT Global Digital  ·  Client: Regencare.in  ·  Competitive Study v2.0 (Extended with Hospital Sites)  ·  May 2026  ·  Confidential")
x.fill=F(NAVY);x.font=Ft(GOLD,False,9,italic=True);x.alignment=Al("center","center",False)

# ═══════════════════════════════════════════════════════════════════════════
out=r"c:\project\AI RESEARCH INTELLIGENCE SYSTEM\Regencare_Competitive_Study_v2.xlsx"
wb.save(out)
print(f"Saved: {out}")
